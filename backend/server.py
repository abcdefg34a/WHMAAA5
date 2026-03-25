from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, status, Request, BackgroundTasks, Header
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, validator
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import random
import string
import re
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.units import cm
import base64
import aiofiles
import math
from collections import defaultdict
import time
from PIL import Image as PILImage
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import csv
import json
import boto3
from botocore.exceptions import ClientError
import pyotp
import qrcode
from qrcode.image.pure import PyPNGImage
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import resend
import asyncio

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Resend Email Configuration
RESEND_API_KEY = os.environ.get('RESEND_API_KEY')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'noreply@werhatmeinautoabgeschleppt.de')
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'info@werhatmeinautoabgeschleppt.de')
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:3000')

# Create directories
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
LOG_DIR = ROOT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
BACKUP_DIR = ROOT_DIR / "backups"
BACKUP_DIR.mkdir(exist_ok=True)

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-super-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Rate Limiting Configuration
RATE_LIMIT_WINDOW = 900  # 15 minutes in seconds
MAX_LOGIN_ATTEMPTS = 5   # Max attempts per window
login_attempts: Dict[str, List[float]] = defaultdict(list)

# Image Compression Settings
MAX_IMAGE_SIZE = (1920, 1080)  # Max dimensions
JPEG_QUALITY = 75  # Quality for JPEG compression

# DSGVO Data Retention Settings (in days)
DSGVO_RETENTION_DAYS = int(os.environ.get('DSGVO_RETENTION_DAYS', 180))  # 6 months default for personal data

# Steuerrechtliche Aufbewahrungsfristen (in Jahren)
# § 147 AO / § 257 HGB: Rechnungen und Buchungsbelege müssen 10 Jahre aufbewahrt werden
INVOICE_RETENTION_YEARS = int(os.environ.get('INVOICE_RETENTION_YEARS', 10))  # 10 years for invoices

# APScheduler instance (will be started in startup_event)
scheduler = AsyncIOScheduler()

# Create the main app
app = FastAPI(title="Abschlepp-Management API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# ==================== STRUCTURED LOGGING ====================
def setup_logging():
    """Setup structured logging with file rotation"""
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter(log_format))
    
    # File handler with rotation (10MB max, keep 5 backups)
    file_handler = RotatingFileHandler(
        LOG_DIR / "app.log",
        maxBytes=10*1024*1024,
        backupCount=5
    )
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(log_format))
    
    # Error file handler
    error_handler = RotatingFileHandler(
        LOG_DIR / "error.log",
        maxBytes=10*1024*1024,
        backupCount=5
    )
    error_handler.setLevel(logging.ERROR)
    error_handler.setFormatter(logging.Formatter(log_format))
    
    # Audit log handler
    audit_handler = RotatingFileHandler(
        LOG_DIR / "audit.log",
        maxBytes=10*1024*1024,
        backupCount=10
    )
    audit_handler.setLevel(logging.INFO)
    audit_handler.setFormatter(logging.Formatter('%(asctime)s - %(message)s'))
    
    # Configure root logger
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.INFO)
    root_logger.addHandler(console_handler)
    root_logger.addHandler(file_handler)
    root_logger.addHandler(error_handler)
    
    # Configure audit logger
    audit_logger = logging.getLogger('audit')
    audit_logger.addHandler(audit_handler)
    audit_logger.setLevel(logging.INFO)
    
    return logging.getLogger(__name__), audit_logger

logger, audit_logger = setup_logging()

# ==================== EMAIL SERVICE (RESEND) ====================

class EmailService:
    """Service for sending emails via Resend"""
    
    def __init__(self):
        """Initialize Resend client"""
        if RESEND_API_KEY:
            resend.api_key = RESEND_API_KEY
            self.enabled = True
            logger.info("Resend Email Service initialized")
        else:
            self.enabled = False
            logger.warning("Resend not configured - emails will be logged only")
        
        self.sender_email = SENDER_EMAIL
        self.admin_email = ADMIN_EMAIL
        self.frontend_url = FRONTEND_URL
    
    async def send_email_async(self, to_email: str, subject: str, html_content: str) -> Optional[str]:
        """Send email via Resend (async/non-blocking)"""
        if not self.enabled:
            logger.info(f"[EMAIL MOCK] To: {to_email}, Subject: {subject}")
            print(f"\n{'='*60}")
            print(f"📧 EMAIL (Resend nicht konfiguriert - nur Log)")
            print(f"An: {to_email}")
            print(f"Betreff: {subject}")
            print(f"{'='*60}\n")
            return None
        
        try:
            params = {
                "from": self.sender_email,
                "to": [to_email],
                "subject": subject,
                "html": html_content
            }
            
            # Run sync SDK in thread to keep FastAPI non-blocking
            email_response = await asyncio.to_thread(resend.Emails.send, params)
            message_id = email_response.get("id") if isinstance(email_response, dict) else getattr(email_response, 'id', None)
            logger.info(f"Email sent successfully to {to_email}, MessageId: {message_id}")
            return message_id
            
        except Exception as e:
            logger.error(f"Failed to send email to {to_email}: {str(e)}")
            raise Exception(f"E-Mail-Versand fehlgeschlagen: {str(e)}")
    
    def send_email(self, to_email: str, subject: str, html_content: str, text_content: str = None) -> Optional[str]:
        """Send email via Resend (sync wrapper for backwards compatibility)"""
        if not self.enabled:
            logger.info(f"[EMAIL MOCK] To: {to_email}, Subject: {subject}")
            print(f"\n{'='*60}")
            print(f"📧 EMAIL (Resend nicht konfiguriert - nur Log)")
            print(f"An: {to_email}")
            print(f"Betreff: {subject}")
            print(f"{'='*60}\n")
            return None
        
        try:
            params = {
                "from": self.sender_email,
                "to": [to_email],
                "subject": subject,
                "html": html_content
            }
            
            email_response = resend.Emails.send(params)
            message_id = email_response.get("id") if isinstance(email_response, dict) else getattr(email_response, 'id', None)
            logger.info(f"Email sent successfully to {to_email}, MessageId: {message_id}")
            return message_id
            
        except Exception as e:
            logger.error(f"Failed to send email to {to_email}: {str(e)}")
            raise Exception(f"E-Mail-Versand fehlgeschlagen: {str(e)}")
    
    def send_password_reset_email(self, to_email: str, reset_token: str, user_name: str) -> Optional[str]:
        """Send password reset email"""
        reset_link = f"{self.frontend_url}/reset-password?token={reset_token}"
        
        subject = "Passwort zurücksetzen - AbschleppPortal"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; }}
                .container {{ background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 40px 30px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .header h1 {{ color: #1e293b; margin: 0; font-size: 24px; }}
                .logo {{ font-size: 32px; margin-bottom: 10px; }}
                .content {{ margin: 20px 0; }}
                .button {{ display: inline-block; padding: 14px 32px; background-color: #f97316; color: white; text-decoration: none; border-radius: 6px; font-weight: 600; margin: 20px 0; }}
                .button:hover {{ background-color: #ea580c; }}
                .link-box {{ background-color: #f1f5f9; padding: 12px; border-radius: 4px; word-break: break-all; font-size: 13px; margin: 15px 0; }}
                .warning {{ background-color: #fef3c7; border: 1px solid #fcd34d; border-radius: 6px; padding: 12px; margin: 20px 0; font-size: 14px; }}
                .footer {{ font-size: 12px; color: #64748b; margin-top: 40px; border-top: 1px solid #e2e8f0; padding-top: 20px; text-align: center; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <div class="logo">🚗</div>
                    <h1>Passwort zurücksetzen</h1>
                </div>
                
                <div class="content">
                    <p>Hallo {user_name},</p>
                    
                    <p>Sie haben eine Anfrage zum Zurücksetzen Ihres Passworts gestellt. Klicken Sie auf den Button unten, um ein neues Passwort zu erstellen:</p>
                    
                    <center>
                        <a href="{reset_link}" class="button">Passwort zurücksetzen</a>
                    </center>
                    
                    <p>Oder kopieren Sie diesen Link in Ihren Browser:</p>
                    <div class="link-box">{reset_link}</div>
                    
                    <div class="warning">
                        <strong>⚠️ Hinweis:</strong> Dieser Link ist 1 Stunde gültig. Falls Sie diese Anfrage nicht gestellt haben, ignorieren Sie diese E-Mail.
                    </div>
                </div>
                
                <div class="footer">
                    <p>Diese E-Mail wurde automatisch gesendet. Bitte antworten Sie nicht direkt darauf.</p>
                    <p>© {datetime.now().year} AbschleppPortal - werhatmeinautoabgeschleppt.de</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        text_content = f"""
Passwort zurücksetzen - AbschleppPortal

Hallo {user_name},

Sie haben eine Anfrage zum Zurücksetzen Ihres Passworts gestellt.

Klicken Sie auf diesen Link, um ein neues Passwort zu erstellen:
{reset_link}

Dieser Link ist 1 Stunde gültig. Falls Sie diese Anfrage nicht gestellt haben, ignorieren Sie diese E-Mail.

Mit freundlichen Grüßen,
Ihr AbschleppPortal Team
        """
        
        return self.send_email(to_email, subject, html_content, text_content)
    
    def send_registration_confirmation(self, to_email: str, user_name: str, role: str) -> Optional[str]:
        """Send registration confirmation email"""
        role_name = "Behörde" if role == "authority" else "Abschleppdienst"
        
        subject = f"Registrierung bestätigt - AbschleppPortal"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; }}
                .container {{ background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 40px 30px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .header h1 {{ color: #16a34a; margin: 0; font-size: 24px; }}
                .success-icon {{ font-size: 48px; margin-bottom: 15px; }}
                .content {{ margin: 20px 0; }}
                .info-box {{ background-color: #dbeafe; border: 1px solid #3b82f6; border-radius: 6px; padding: 15px; margin: 20px 0; }}
                .footer {{ font-size: 12px; color: #64748b; margin-top: 40px; border-top: 1px solid #e2e8f0; padding-top: 20px; text-align: center; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <div class="success-icon">✅</div>
                    <h1>Willkommen bei AbschleppPortal!</h1>
                </div>
                
                <div class="content">
                    <p>Hallo {user_name},</p>
                    
                    <p>Vielen Dank für Ihre Registrierung als <strong>{role_name}</strong>!</p>
                    
                    <div class="info-box">
                        <strong>ℹ️ Nächste Schritte:</strong><br>
                        Ihre Registrierung muss von einem Administrator freigegeben werden, bevor Sie das Portal nutzen können. Sie erhalten eine weitere E-Mail, sobald Ihr Konto aktiviert wurde.
                    </div>
                    
                    <p>Bei Fragen stehen wir Ihnen gerne zur Verfügung.</p>
                </div>
                
                <div class="footer">
                    <p>© {datetime.now().year} AbschleppPortal - werhatmeinautoabgeschleppt.de</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return self.send_email(to_email, subject, html_content)
    
    def send_account_approved_email(self, to_email: str, user_name: str, role: str) -> Optional[str]:
        """Send account approval notification"""
        role_name = "Behörde" if role == "authority" else "Abschleppdienst"
        login_url = f"{self.frontend_url}/portal"
        
        subject = "Ihr Konto wurde freigeschaltet - AbschleppPortal"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; }}
                .container {{ background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 40px 30px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .header h1 {{ color: #16a34a; margin: 0; font-size: 24px; }}
                .success-icon {{ font-size: 48px; margin-bottom: 15px; }}
                .button {{ display: inline-block; padding: 14px 32px; background-color: #16a34a; color: white; text-decoration: none; border-radius: 6px; font-weight: 600; margin: 20px 0; }}
                .footer {{ font-size: 12px; color: #64748b; margin-top: 40px; border-top: 1px solid #e2e8f0; padding-top: 20px; text-align: center; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <div class="success-icon">🎉</div>
                    <h1>Konto freigeschaltet!</h1>
                </div>
                
                <div class="content">
                    <p>Hallo {user_name},</p>
                    
                    <p>Gute Nachrichten! Ihr Konto als <strong>{role_name}</strong> wurde von einem Administrator freigegeben.</p>
                    
                    <p>Sie können sich jetzt anmelden und das Portal nutzen:</p>
                    
                    <center>
                        <a href="{login_url}" class="button">Zum Portal anmelden</a>
                    </center>
                </div>
                
                <div class="footer">
                    <p>© {datetime.now().year} AbschleppPortal - werhatmeinautoabgeschleppt.de</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return self.send_email(to_email, subject, html_content)
    
    def send_job_notification(self, to_email: str, job_number: str, license_plate: str, status: str, service_name: str = None) -> Optional[str]:
        """Send job status notification"""
        status_labels = {
            'assigned': 'Zugewiesen',
            'on_site': 'Abschleppdienst vor Ort',
            'towed': 'Fahrzeug abgeschleppt',
            'in_yard': 'Fahrzeug im Hof',
            'delivered_to_authority': 'An Behörde übergeben',
            'released': 'Fahrzeug abgeholt'
        }
        status_label = status_labels.get(status, status)
        
        subject = f"Auftrag {job_number} - {status_label}"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; }}
                .container {{ background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 40px 30px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .header h1 {{ color: #1e293b; margin: 0; font-size: 24px; }}
                .status-badge {{ display: inline-block; padding: 8px 16px; background-color: #f97316; color: white; border-radius: 20px; font-weight: 600; }}
                .details {{ background-color: #f1f5f9; border-radius: 6px; padding: 15px; margin: 20px 0; }}
                .detail-row {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #e2e8f0; }}
                .footer {{ font-size: 12px; color: #64748b; margin-top: 40px; border-top: 1px solid #e2e8f0; padding-top: 20px; text-align: center; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🚗 Auftrags-Update</h1>
                </div>
                
                <div class="content">
                    <center>
                        <span class="status-badge">{status_label}</span>
                    </center>
                    
                    <div class="details">
                        <div class="detail-row">
                            <strong>Auftragsnummer:</strong>
                            <span>{job_number}</span>
                        </div>
                        <div class="detail-row">
                            <strong>Kennzeichen:</strong>
                            <span>{license_plate}</span>
                        </div>
                        {f'<div class="detail-row"><strong>Abschleppdienst:</strong><span>{service_name}</span></div>' if service_name else ''}
                    </div>
                </div>
                
                <div class="footer">
                    <p>© {datetime.now().year} AbschleppPortal - werhatmeinautoabgeschleppt.de</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return self.send_email(to_email, subject, html_content)
    
    # ==================== BACKUP NOTIFICATION METHODS ====================
    
    async def send_backup_failure_alert(self, backup_type: str, error_message: str, backup_id: str = None) -> Optional[str]:
        """Send alert email when a backup fails"""
        subject = f"⚠️ Backup-Fehler - {backup_type.upper()}"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; }}
                .container {{ background-color: #fef2f2; border: 2px solid #ef4444; border-radius: 8px; padding: 30px; }}
                .header {{ text-align: center; margin-bottom: 20px; }}
                .header h1 {{ color: #dc2626; margin: 0; font-size: 22px; }}
                .alert-icon {{ font-size: 48px; margin-bottom: 15px; }}
                .error-box {{ background-color: #fee2e2; border: 1px solid #fca5a5; border-radius: 6px; padding: 15px; margin: 20px 0; font-family: monospace; font-size: 13px; white-space: pre-wrap; word-break: break-word; }}
                .details {{ background-color: #fff; border-radius: 6px; padding: 15px; margin: 20px 0; }}
                .detail-row {{ padding: 8px 0; border-bottom: 1px solid #f3f4f6; }}
                .detail-row:last-child {{ border-bottom: none; }}
                .action-box {{ background-color: #fef3c7; border: 1px solid #fcd34d; border-radius: 6px; padding: 15px; margin: 20px 0; }}
                .footer {{ font-size: 12px; color: #64748b; margin-top: 30px; text-align: center; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <div class="alert-icon">🚨</div>
                    <h1>Backup fehlgeschlagen!</h1>
                </div>
                
                <div class="details">
                    <div class="detail-row">
                        <strong>Backup-Typ:</strong> {backup_type}
                    </div>
                    <div class="detail-row">
                        <strong>Zeitpunkt:</strong> {datetime.now(timezone.utc).strftime('%d.%m.%Y um %H:%M:%S Uhr')} UTC
                    </div>
                    {f'<div class="detail-row"><strong>Backup-ID:</strong> {backup_id}</div>' if backup_id else ''}
                </div>
                
                <p><strong>Fehlermeldung:</strong></p>
                <div class="error-box">{error_message}</div>
                
                <div class="action-box">
                    <strong>⚡ Empfohlene Maßnahmen:</strong>
                    <ul style="margin: 10px 0 0 0; padding-left: 20px;">
                        <li>Prüfen Sie die Backend-Logs auf weitere Details</li>
                        <li>Stellen Sie sicher, dass genügend Speicherplatz vorhanden ist</li>
                        <li>Überprüfen Sie die Datenbankverbindung</li>
                        <li>Führen Sie ein manuelles Backup durch, um das Problem zu isolieren</li>
                    </ul>
                </div>
                
                <div class="footer">
                    <p>Diese automatische Benachrichtigung wurde vom AbschleppPortal Backup-System gesendet.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return await self.send_email_async(self.admin_email, subject, html_content)
    
    async def send_weekly_backup_report(self, stats: dict) -> Optional[str]:
        """Send weekly backup status report"""
        subject = f"📊 Wöchentlicher Backup-Report - KW {datetime.now(timezone.utc).isocalendar()[1]}"
        
        # Calculate health status
        total_backups = stats.get('total_backups', 0)
        failed_backups = stats.get('failed_backups', 0)
        success_rate = ((total_backups - failed_backups) / total_backups * 100) if total_backups > 0 else 0
        
        health_color = '#16a34a' if success_rate >= 95 else '#f59e0b' if success_rate >= 80 else '#dc2626'
        health_text = 'Ausgezeichnet' if success_rate >= 95 else 'Gut' if success_rate >= 80 else 'Kritisch'
        health_icon = '✅' if success_rate >= 95 else '⚠️' if success_rate >= 80 else '❌'
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; }}
                .container {{ background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 30px; }}
                .header {{ text-align: center; margin-bottom: 25px; }}
                .header h1 {{ color: #1e293b; margin: 0; font-size: 22px; }}
                .health-badge {{ display: inline-block; padding: 10px 20px; background-color: {health_color}; color: white; border-radius: 25px; font-weight: 600; font-size: 16px; margin: 15px 0; }}
                .stats-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin: 20px 0; }}
                .stat-box {{ background-color: #fff; border: 1px solid #e2e8f0; border-radius: 8px; padding: 15px; text-align: center; }}
                .stat-value {{ font-size: 28px; font-weight: 700; color: #1e293b; }}
                .stat-label {{ font-size: 12px; color: #64748b; text-transform: uppercase; letter-spacing: 0.5px; }}
                .details {{ background-color: #fff; border-radius: 6px; padding: 15px; margin: 20px 0; border: 1px solid #e2e8f0; }}
                .detail-row {{ display: flex; justify-content: space-between; padding: 10px 0; border-bottom: 1px solid #f3f4f6; }}
                .detail-row:last-child {{ border-bottom: none; }}
                .footer {{ font-size: 12px; color: #64748b; margin-top: 30px; text-align: center; border-top: 1px solid #e2e8f0; padding-top: 20px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📊 Wöchentlicher Backup-Report</h1>
                    <p style="color: #64748b; margin: 5px 0;">Kalenderwoche {datetime.now(timezone.utc).isocalendar()[1]} / {datetime.now(timezone.utc).year}</p>
                    <span class="health-badge">{health_icon} {health_text} ({success_rate:.1f}%)</span>
                </div>
                
                <div class="stats-grid">
                    <div class="stat-box">
                        <div class="stat-value">{total_backups}</div>
                        <div class="stat-label">Backups gesamt</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value" style="color: #16a34a;">{total_backups - failed_backups}</div>
                        <div class="stat-label">Erfolgreich</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value" style="color: {'#dc2626' if failed_backups > 0 else '#64748b'};">{failed_backups}</div>
                        <div class="stat-label">Fehlgeschlagen</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{stats.get('total_size_mb', 0):.1f}</div>
                        <div class="stat-label">Speicher (MB)</div>
                    </div>
                </div>
                
                <div class="details">
                    <h3 style="margin-top: 0; color: #1e293b;">📋 Details</h3>
                    <div class="detail-row">
                        <span>Datenbank-Backups:</span>
                        <strong>{stats.get('db_backups', 0)}</strong>
                    </div>
                    <div class="detail-row">
                        <span>Storage-Backups:</span>
                        <strong>{stats.get('storage_backups', 0)}</strong>
                    </div>
                    <div class="detail-row">
                        <span>Cloud-Uploads (Supabase):</span>
                        <strong>{stats.get('cloud_uploads', 0)}</strong>
                    </div>
                    <div class="detail-row">
                        <span>Letztes erfolgreiches Backup:</span>
                        <strong>{stats.get('last_successful', 'N/A')}</strong>
                    </div>
                    <div class="detail-row">
                        <span>Ältestes verfügbares Backup:</span>
                        <strong>{stats.get('oldest_backup', 'N/A')}</strong>
                    </div>
                </div>
                
                <div class="footer">
                    <p>Dieser Report wird jeden Montag automatisch gesendet.</p>
                    <p>© {datetime.now().year} AbschleppPortal - werhatmeinautoabgeschleppt.de</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return await self.send_email_async(self.admin_email, subject, html_content)

# Initialize email service
email_service = EmailService()

# ==================== MODELS ====================

class UserRole:
    ADMIN = "admin"
    AUTHORITY = "authority"
    TOWING_SERVICE = "towing_service"

class ApprovalStatus:
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"

class JobStatus:
    PENDING = "pending"
    ASSIGNED = "assigned"
    ON_SITE = "on_site"
    TOWED = "towed"
    IN_YARD = "in_yard"
    # NEW: Delivered to authority yard (towing service job complete)
    DELIVERED_TO_AUTHORITY = "delivered_to_authority"
    RELEASED = "released"

# Auth Models
class UserRegister(BaseModel):
    email: EmailStr
    password: str
    role: str
    name: str
    # Authority fields
    authority_name: Optional[str] = None
    department: Optional[str] = None
    # Towing service fields
    company_name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    yard_address: Optional[str] = None
    opening_hours: Optional[str] = None
    # NEW: Cost fields
    tow_cost: Optional[float] = None  # Anfahrtskosten
    daily_cost: Optional[float] = None  # Standkosten pro Tag
    # NEW: Business license (Base64)
    business_license: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserLogin2FA(BaseModel):
    temp_token: str
    totp_code: str

class TOTPVerifyRequest(BaseModel):
    totp_code: str

class UserResponse(BaseModel):
    id: str
    email: str
    role: str
    name: str
    authority_name: Optional[str] = None
    department: Optional[str] = None
    company_name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    yard_address: Optional[str] = None
    yard_lat: Optional[float] = None  # NEW: Yard coordinates
    yard_lng: Optional[float] = None  # NEW: Yard coordinates
    opening_hours: Optional[str] = None
    service_code: Optional[str] = None
    linked_services: Optional[List[str]] = None
    linked_authorities: Optional[List[str]] = None  # NEW: Authorities that linked this service
    created_at: Optional[str] = None
    # Pricing fields
    tow_cost: Optional[float] = None
    daily_cost: Optional[float] = None
    business_license: Optional[str] = None
    approval_status: Optional[str] = None
    rejection_reason: Optional[str] = None
    is_blocked: Optional[bool] = None
    # Employee/Dienstnummer fields
    dienstnummer: Optional[str] = None
    parent_authority_id: Optional[str] = None
    is_main_authority: Optional[bool] = None
    # 2FA
    totp_enabled: Optional[bool] = False
    # NEW: Extended pricing settings (for towing services)
    time_based_enabled: Optional[bool] = None
    first_half_hour: Optional[float] = None
    additional_half_hour: Optional[float] = None
    processing_fee: Optional[float] = None
    empty_trip_fee: Optional[float] = None
    night_surcharge: Optional[float] = None
    weekend_surcharge: Optional[float] = None
    heavy_vehicle_surcharge: Optional[float] = None
    # NEW: Flexible weight categories (for towing services)
    weight_categories: Optional[List[dict]] = None
    # NEW: Authority yard model and pricing (for authorities)
    yard_model: Optional[str] = None  # "authority_yard" or "service_yard"
    price_categories: Optional[List[dict]] = None  # Authority's own price categories

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class UpdateCostsRequest(BaseModel):
    tow_cost: float
    daily_cost: float

# NEW: Fahrzeugkategorien mit Preisen (Hamburg-Modell)
class VehicleCategoryPrice(BaseModel):
    id: Optional[str] = None
    name: str  # z.B. "PKW bis 4t"
    description: Optional[str] = None  # z.B. "Kraftfahrzeug mit einer zulässigen Gesamtmasse bis 4t"
    base_price: float  # Grundpreis für erste 24h
    daily_rate: float  # Preis pro weitere 24h
    is_active: bool = True

class VehicleCategoryPriceList(BaseModel):
    categories: List[VehicleCategoryPrice]

# NEW: Weight category model for flexible pricing
class WeightCategory(BaseModel):
    id: Optional[str] = None
    name: str  # z.B. "PKW bis 3,5t", "LKW 3,5-7,5t"
    min_weight: Optional[float] = None  # Mindestgewicht in Tonnen (None = kein Minimum)
    max_weight: Optional[float] = None  # Maximalgewicht in Tonnen (None = kein Maximum)
    surcharge: float = 0  # Zuschlag in Euro
    is_default: bool = False  # Standard-Kategorie (kein Zuschlag)

# NEW: Extended pricing settings for towing services
class PricingSettingsRequest(BaseModel):
    # Zeitbasiert
    time_based_enabled: Optional[bool] = False
    first_half_hour: Optional[float] = None
    additional_half_hour: Optional[float] = None
    # Standard
    tow_cost: Optional[float] = None
    daily_cost: Optional[float] = None
    # Zusatzkosten
    processing_fee: Optional[float] = None  # Bearbeitungsgebühr
    empty_trip_fee: Optional[float] = None  # Leerfahrt
    night_surcharge: Optional[float] = None  # Nachtzuschlag
    weekend_surcharge: Optional[float] = None  # Wochenendzuschlag
    heavy_vehicle_surcharge: Optional[float] = None  # DEPRECATED: Use weight_categories instead
    # NEW: Flexible Gewichtskategorien
    weight_categories: Optional[List[WeightCategory]] = None

class ApproveServiceRequest(BaseModel):
    approved: bool
    rejection_reason: Optional[str] = None

# Authority Employee Models
class CreateEmployeeRequest(BaseModel):
    email: EmailStr
    password: str
    name: str

class EmployeeResponse(BaseModel):
    id: str
    email: str
    name: str
    dienstnummer: str
    is_blocked: Optional[bool] = None
    created_at: str

# Admin User Management Models
class AdminUpdatePasswordRequest(BaseModel):
    new_password: str

class AdminBlockUserRequest(BaseModel):
    blocked: bool

# Job Models
class JobCreate(BaseModel):
    license_plate: Optional[str] = None
    vin: Optional[str] = None
    tow_reason: str
    location_address: str
    location_lat: float
    location_lng: float
    photos: List[str] = []
    notes: Optional[str] = None
    assigned_service_id: Optional[str] = None
    # NEW: For towing service creating jobs on behalf of authority
    for_authority_id: Optional[str] = None
    # NEW: Job type and Sicherstellung fields
    job_type: Optional[str] = "towing"  # "towing" or "sicherstellung"
    # Sicherstellung-specific fields
    sicherstellung_reason: Optional[str] = None
    vehicle_category: Optional[str] = None  # "under_3_5t" or "over_3_5t"
    vehicle_category_id: Optional[str] = None  # Reference to vehicle category for dynamic pricing
    # NEW: Target yard and pricing model
    target_yard: Optional[str] = None  # "authority_yard" or "service_yard"
    # Authority yard location (when target_yard = "authority_yard")
    authority_yard_id: Optional[str] = None
    authority_yard_name: Optional[str] = None
    authority_yard_address: Optional[str] = None
    authority_yard_lat: Optional[float] = None
    authority_yard_lng: Optional[float] = None
    authority_yard_phone: Optional[str] = None
    # Authority price category (when target_yard = "authority_yard")
    authority_price_category_id: Optional[str] = None
    authority_price_category_name: Optional[str] = None
    authority_base_price: Optional[float] = None
    authority_daily_rate: Optional[float] = None
    # NEW: Weight category for flexible pricing (from towing service, when target_yard = "service_yard")
    weight_category_id: Optional[str] = None
    weight_category_name: Optional[str] = None
    weight_category_surcharge: Optional[float] = None
    ordering_authority: Optional[str] = None
    contact_attempts: Optional[bool] = None
    contact_attempts_notes: Optional[str] = None
    estimated_vehicle_value: Optional[float] = None

class JobUpdate(BaseModel):
    status: Optional[str] = None
    photos: Optional[List[str]] = None
    notes: Optional[str] = None
    service_notes: Optional[str] = None
    # Release info
    owner_first_name: Optional[str] = None
    owner_last_name: Optional[str] = None
    owner_address: Optional[str] = None
    payment_method: Optional[str] = None
    payment_amount: Optional[float] = None
    # NEW: Empty trip flag
    is_empty_trip: Optional[bool] = None

# NEW: Model for editing job vehicle data (Kennzeichen, FIN, etc.)
class JobEditData(BaseModel):
    license_plate: Optional[str] = None
    vin: Optional[str] = None
    tow_reason: Optional[str] = None
    notes: Optional[str] = None
    location_address: Optional[str] = None
    location_lat: Optional[float] = None
    location_lng: Optional[float] = None

class BulkStatusUpdate(BaseModel):
    job_ids: List[str]
    status: str

class JobResponse(BaseModel):
    id: str
    job_number: str
    license_plate: Optional[str] = None
    vin: Optional[str] = None
    tow_reason: str
    location_address: Optional[str] = None
    location_lat: Optional[float] = None
    location_lng: Optional[float] = None
    photos: List[str] = []
    notes: Optional[str] = None
    status: str
    created_by_id: Optional[str] = None
    created_by_name: Optional[str] = None
    created_by_authority: Optional[str] = None
    created_by_dienstnummer: Optional[str] = None
    authority_id: Optional[str] = None
    assigned_service_id: Optional[str] = None
    assigned_service_name: Optional[str] = None
    service_notes: Optional[str] = None
    service_photos: List[str] = []
    owner_first_name: Optional[str] = None
    owner_last_name: Optional[str] = None
    owner_address: Optional[str] = None
    payment_method: Optional[str] = None
    payment_amount: Optional[float] = None
    created_at: str
    updated_at: str
    on_site_at: Optional[str] = None
    towed_at: Optional[str] = None
    in_yard_at: Optional[str] = None
    # NEW: When vehicle was delivered to authority yard
    delivered_to_authority_at: Optional[str] = None
    released_at: Optional[str] = None
    accepted_at: Optional[str] = None  # NEW: When job was accepted
    # NEW: Job type and Sicherstellung fields
    job_type: Optional[str] = "towing"
    sicherstellung_reason: Optional[str] = None
    vehicle_category: Optional[str] = None
    vehicle_category_id: Optional[str] = None  # Referenz zur Preiskategorie
    # NEW: Target yard model (authority_yard or service_yard)
    target_yard: Optional[str] = None
    # Authority yard location (when target_yard = "authority_yard")
    authority_yard_id: Optional[str] = None
    authority_yard_name: Optional[str] = None
    authority_yard_address: Optional[str] = None
    authority_yard_lat: Optional[float] = None
    authority_yard_lng: Optional[float] = None
    authority_yard_phone: Optional[str] = None
    # Authority pricing (when target_yard = "authority_yard")
    authority_price_category_id: Optional[str] = None
    authority_price_category_name: Optional[str] = None
    authority_base_price: Optional[float] = None
    authority_daily_rate: Optional[float] = None
    # NEW: Weight category for flexible pricing (service_yard)
    weight_category_id: Optional[str] = None
    weight_category_name: Optional[str] = None
    weight_category_surcharge: Optional[float] = None
    ordering_authority: Optional[str] = None
    contact_attempts: Optional[bool] = None
    contact_attempts_notes: Optional[str] = None
    estimated_vehicle_value: Optional[float] = None
    is_empty_trip: Optional[bool] = None
    # NEW: Calculated costs breakdown
    calculated_costs: Optional[dict] = None
    # NEW: Track if created by towing service
    created_by_service: Optional[bool] = False
    # NEW: Service invoice (when authority releases from their yard)
    service_invoice_amount: Optional[float] = None
    service_invoice_created_at: Optional[str] = None
    # DSGVO anonymization flags
    anonymized: Optional[bool] = None
    personal_data_anonymized: Optional[bool] = None
    personal_data_anonymized_at: Optional[str] = None

class VehicleSearchResult(BaseModel):
    found: bool
    job_number: Optional[str] = None
    license_plate: Optional[str] = None
    status: Optional[str] = None
    towed_at: Optional[str] = None
    in_yard_at: Optional[str] = None
    yard_address: Optional[str] = None
    yard_lat: Optional[float] = None  # NEW: Yard coordinates for map
    yard_lng: Optional[float] = None  # NEW: Yard coordinates for map
    company_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    opening_hours: Optional[str] = None
    # Cost calculation
    tow_cost: Optional[float] = None
    daily_cost: Optional[float] = None
    days_in_yard: Optional[int] = None
    processing_fee: Optional[float] = None  # Bearbeitungsgebühr
    empty_trip_fee: Optional[float] = None  # Leerfahrt
    night_surcharge: Optional[float] = None  # Nachtzuschlag
    weekend_surcharge: Optional[float] = None  # Wochenendzuschlag
    heavy_vehicle_surcharge: Optional[float] = None  # Schwerlast ab 3,5t
    total_cost: Optional[float] = None
    # Location coordinates for map - REMOVED, using yard coordinates instead
    location_lat: Optional[float] = None
    location_lng: Optional[float] = None
    # NEW: Additional info for towed status
    tow_reason: Optional[str] = None
    location_address: Optional[str] = None
    created_by_authority: Optional[str] = None

class LinkServiceRequest(BaseModel):
    service_code: str

class StatsResponse(BaseModel):
    total_jobs: int
    pending_jobs: int
    in_yard_jobs: int
    released_jobs: int
    total_services: int
    total_authorities: int
    pending_approvals: int

# ==================== HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def validate_password(password: str) -> tuple[bool, str]:
    """
    Validate password strength:
    - Minimum 8 characters
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one number
    Returns (is_valid, error_message)
    """
    if len(password) < 8:
        return False, "Passwort muss mindestens 8 Zeichen lang sein"
    if not re.search(r'[A-Z]', password):
        return False, "Passwort muss mindestens einen Großbuchstaben enthalten"
    if not re.search(r'[a-z]', password):
        return False, "Passwort muss mindestens einen Kleinbuchstaben enthalten"
    if not re.search(r'\d', password):
        return False, "Passwort muss mindestens eine Zahl enthalten"
    return True, ""

def check_rate_limit(identifier: str) -> tuple[bool, int]:
    """
    Check if login attempts are within rate limit.
    Returns (is_allowed, seconds_until_reset)
    """
    current_time = time.time()
    
    # NEW: Global cleanup to prevent memory exhaustion by botnets
    if len(login_attempts) > 10000:
        keys_to_delete = []
        for k, attempts in list(login_attempts.items()):
            valid_attempts = [t for t in attempts if current_time - t < RATE_LIMIT_WINDOW]
            if not valid_attempts:
                keys_to_delete.append(k)
            else:
                login_attempts[k] = valid_attempts
        for k in keys_to_delete:
            login_attempts.pop(k, None)

    # Clean old attempts for this specific user/IP
    login_attempts[identifier] = [
        t for t in login_attempts[identifier] 
        if current_time - t < RATE_LIMIT_WINDOW
    ]
    
    if len(login_attempts[identifier]) >= MAX_LOGIN_ATTEMPTS:
        oldest_attempt = min(login_attempts[identifier])
        seconds_remaining = int(RATE_LIMIT_WINDOW - (current_time - oldest_attempt))
        return False, seconds_remaining
        
    if len(login_attempts[identifier]) == 0:
        login_attempts.pop(identifier, None)
    
    return True, 0

def record_login_attempt(identifier: str):
    """Record a failed login attempt"""
    login_attempts[identifier].append(time.time())

def clear_login_attempts(identifier: str):
    """Clear login attempts after successful login"""
    if identifier in login_attempts:
        del login_attempts[identifier]

def generate_reset_token() -> str:
    """Generate a secure password reset token"""
    return ''.join(random.choices(string.ascii_letters + string.digits, k=64))

# ==================== IMAGE COMPRESSION ====================

def compress_image_base64(base64_data: str, max_size: tuple = MAX_IMAGE_SIZE, quality: int = JPEG_QUALITY) -> str:
    """
    Compress a base64 encoded image.
    Returns compressed base64 string.
    """
    try:
        # Handle data URL format
        if ',' in base64_data:
            header, data = base64_data.split(',', 1)
        else:
            header = "data:image/jpeg;base64"
            data = base64_data
        
        # Decode base64
        image_data = base64.b64decode(data)
        
        # Open with Pillow
        img = PILImage.open(BytesIO(image_data))
        
        # Convert RGBA to RGB if needed
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # Resize if too large
        if img.size[0] > max_size[0] or img.size[1] > max_size[1]:
            img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
        
        # Save compressed
        output = BytesIO()
        img.save(output, format='JPEG', quality=quality, optimize=True)
        output.seek(0)
        
        # Encode back to base64
        compressed_data = base64.b64encode(output.read()).decode()
        
        logger.info(f"Image compressed: {len(data)} -> {len(compressed_data)} bytes ({int((1-len(compressed_data)/len(data))*100)}% reduction)")
        
        return f"data:image/jpeg;base64,{compressed_data}"
    except Exception as e:
        logger.error(f"Image compression failed: {e}")
        return base64_data  # Return original if compression fails

# ==================== AUDIT LOGGING ====================

async def log_audit(action: str, user_id: str, user_name: str, details: Dict[str, Any] = None):
    """Log an audit event to both file and database"""
    now = datetime.now(timezone.utc).isoformat()
    
    audit_entry = {
        "id": str(uuid.uuid4()),
        "timestamp": now,
        "action": action,
        "user_id": user_id,
        "user_name": user_name,
        "details": details or {},
        "created_at": now
    }
    
    # Log to file
    audit_logger.info(json.dumps(audit_entry, ensure_ascii=False))
    
    # Log to database
    try:
        await db.audit_logs.insert_one(audit_entry)
    except Exception as e:
        logger.error(f"Failed to save audit log to DB: {e}")
    
    return audit_entry

def create_token(user_id: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def generate_service_code() -> str:
    letters = ''.join(random.choices(string.ascii_letters, k=4))
    numbers = ''.join(random.choices(string.digits, k=2))
    return letters + numbers

def generate_job_number() -> str:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    random_part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"TOW-{timestamp}-{random_part}"

async def generate_dienstnummer(authority_id: str) -> str:
    """Generate unique Dienstnummer for authority employees"""
    # Count existing employees for this authority
    count = await db.users.count_documents({
        "$or": [
            {"id": authority_id},
            {"parent_authority_id": authority_id}
        ],
        "dienstnummer": {"$exists": True}
    })
    # Format: DN-XXXX-NNN (DN = Dienstnummer, XXXX = first 4 chars of authority_id, NNN = sequential number)
    prefix = authority_id[:4].upper()
    return f"DN-{prefix}-{str(count + 1).zfill(3)}"

def calculate_days_in_yard(in_yard_at: str) -> int:
    """Calculate number of days a vehicle has been in the yard"""
    if not in_yard_at:
        return 0
    try:
        in_yard_date = datetime.fromisoformat(in_yard_at.replace('Z', '+00:00'))
        now = datetime.now(timezone.utc)
        delta = now - in_yard_date
        return max(1, math.ceil(delta.total_seconds() / 86400))  # Minimum 1 day
    except:
        return 1

def calculate_total_cost(tow_cost: float, daily_cost: float, days: int) -> float:
    """Calculate total cost: tow_cost + (daily_cost * days)"""
    return (tow_cost or 0) + (daily_cost or 0) * days

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        # Try to find user by UUID id field first, then by MongoDB _id
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user:
            # Try by MongoDB ObjectId (legacy support)
            try:
                from bson.objectid import ObjectId
                user = await db.users.find_one({"_id": ObjectId(user_id)}, {"password": 0})
                if user:
                    user["id"] = str(user.pop("_id"))
            except:
                pass
        
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_role(roles: List[str], user: dict = Depends(get_current_user)):
    if user["role"] not in roles:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return user

# ==================== AUTH ROUTES ====================

class PasswordResetRequest(BaseModel):
    email: EmailStr

class PasswordResetConfirm(BaseModel):
    token: str
    new_password: str

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(data: UserRegister, request: Request, background_tasks: BackgroundTasks):
    # Check if email exists
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="E-Mail bereits registriert")
    
    # Validate password strength
    is_valid, error_msg = validate_password(data.password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    user_doc = {
        "id": user_id,
        "email": data.email,
        "password": hash_password(data.password),
        "role": data.role,
        "name": data.name,
        "created_at": now,
        "email_verified": False  # Email verification status
    }
    
    if data.role == UserRole.AUTHORITY:
        user_doc["authority_name"] = data.authority_name
        user_doc["department"] = data.department
        user_doc["linked_services"] = []
        user_doc["is_main_authority"] = True
        user_doc["parent_authority_id"] = None
        # Generate Dienstnummer for main authority
        user_doc["dienstnummer"] = await generate_dienstnummer(user_id)
        # Authority also needs approval
        user_doc["approval_status"] = ApprovalStatus.PENDING
        user_doc["rejection_reason"] = None
    elif data.role == UserRole.TOWING_SERVICE:
        # Check if business license is provided
        if not data.business_license:
            raise HTTPException(status_code=400, detail="Gewerbenachweis ist erforderlich")
        
        user_doc["company_name"] = data.company_name
        user_doc["phone"] = data.phone
        user_doc["address"] = data.address
        user_doc["yard_address"] = data.yard_address
        user_doc["opening_hours"] = data.opening_hours
        user_doc["service_code"] = generate_service_code()
        # NEW: Cost fields
        user_doc["tow_cost"] = data.tow_cost or 0
        user_doc["daily_cost"] = data.daily_cost or 0
        # NEW: Business license and approval status
        user_doc["business_license"] = data.business_license
        user_doc["approval_status"] = ApprovalStatus.PENDING
        user_doc["rejection_reason"] = None
    
    await db.users.insert_one(user_doc)
    
    # Audit log new registration
    await log_audit("USER_REGISTERED", user_id, data.name, {
        "email": data.email,
        "role": data.role,
        "approval_required": data.role in [UserRole.TOWING_SERVICE, UserRole.AUTHORITY]
    })
    
    # Send registration confirmation email via BackgroundTasks
    try:
        background_tasks.add_task(
            email_service.send_registration_confirmation,
            to_email=data.email,
            user_name=data.name,
            role=data.role
        )
        logger.info(f"Registration confirmation email queued for {data.email}")
    except Exception as e:
        logger.error(f"Failed to send registration email: {e}")
    
    # For towing services and authorities: Don't return token, they need approval first
    if data.role in [UserRole.TOWING_SERVICE, UserRole.AUTHORITY]:
        raise HTTPException(
            status_code=202,  # Accepted but not complete
            detail="Registrierung erfolgreich! Ihr Konto muss erst von einem Administrator freigeschaltet werden. Sie erhalten eine Benachrichtigung, sobald Ihr Konto aktiviert wurde."
        )
    
    # For admin and other roles: Return token immediately
    token = create_token(user_id, data.role)
    user_doc.pop("password")
    user_doc.pop("_id", None)
    
    return TokenResponse(access_token=token, user=UserResponse(**user_doc))

@api_router.post("/auth/login")
async def login(data: UserLogin, request: Request):
    try:
        # Get client IP for rate limiting
        client_ip = request.client.host if request.client else "unknown"
        rate_limit_key = f"{client_ip}:{data.email}"
        
        # Check rate limit
        is_allowed, seconds_remaining = check_rate_limit(rate_limit_key)
        if not is_allowed:
            minutes = seconds_remaining // 60
            raise HTTPException(
                status_code=429, 
                detail=f"Zu viele Anmeldeversuche. Bitte warten Sie {minutes} Minuten."
            )
        
        user = await db.users.find_one({"email": data.email})
        if not user or not verify_password(data.password, user["password"]):
            # Record failed attempt
            record_login_attempt(rate_limit_key)
            attempts_left = MAX_LOGIN_ATTEMPTS - len(login_attempts[rate_limit_key])
            # Audit log failed login attempt
            await log_audit("LOGIN_FAILED", "unknown", data.email, {
                "email": data.email,
                "ip_address": client_ip,
                "reason": "invalid_credentials"
            })
            raise HTTPException(
                status_code=401, 
                detail=f"Ungültige Anmeldedaten. Noch {attempts_left} Versuche übrig."
            )
        
        # Check if user is blocked
        if user.get("is_blocked"):
            raise HTTPException(status_code=403, detail="Ihr Konto wurde gesperrt. Bitte kontaktieren Sie den Administrator.")
        
        # Check if authority or towing service is approved (employees don't need approval)
        if user["role"] in [UserRole.TOWING_SERVICE, UserRole.AUTHORITY]:
            # Skip approval check for employee accounts (they inherit from parent)
            if user.get("is_main_authority") == False and user.get("parent_authority_id"):
                pass  # Employee accounts don't need separate approval
            elif user["role"] == UserRole.AUTHORITY and not user.get("is_main_authority"):
                pass  # Non-main authority employees
            else:
                # Main accounts need approval
                approval_status = user.get("approval_status")
                if approval_status is None or approval_status == ApprovalStatus.PENDING:
                    raise HTTPException(status_code=403, detail="Ihr Konto wartet noch auf Freischaltung durch einen Administrator")
                elif approval_status == ApprovalStatus.REJECTED:
                    rejection_reason = user.get("rejection_reason", "")
                    raise HTTPException(status_code=403, detail=f"Ihre Registrierung wurde abgelehnt: {rejection_reason}. Sie können sich erneut registrieren.")
        
        # Clear rate limit on successful login info fetch
        clear_login_attempts(rate_limit_key)
    
        # Use existing id field or fall back to MongoDB _id
        if not user.get("id"):
            user["id"] = str(user["_id"])
        
        # 2FA CHECK
        if user.get("totp_enabled"):
            # Create a temp token valid for 5 minutes just for 2FA validation
            temp_payload = {
                "user_id": user["id"],
                "type": "2fa_temp",
                "exp": datetime.now(timezone.utc) + timedelta(minutes=5)
            }
            temp_token = jwt.encode(temp_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
            return {
                "requires_2fa": True,
                "temp_token": temp_token
            }
            
        # Standard Login Flow Without 2FA
        await log_audit("USER_LOGIN", user["id"], user.get("name", user["email"]), {
            "email": user["email"],
            "role": user["role"],
            "ip_address": client_ip,
            "2fa_used": False
        })
        
        token = create_token(user["id"], user["role"])
        user.pop("password")
        user.pop("_id", None)
        # Don't send secret to frontend
        user.pop("totp_secret", None)
        
        return TokenResponse(access_token=token, user=UserResponse(**user))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal Server Error")

@api_router.post("/auth/login/2fa", response_model=TokenResponse)
async def login_2fa(data: UserLogin2FA, request: Request):
    # Get client IP
    client_ip = request.client.host if request.client else "unknown"
    
    try:
        # Verify temp token
        payload = jwt.decode(data.temp_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "2fa_temp":
            raise HTTPException(status_code=401, detail="Ungültiger Token-Typ")
            
        user_id = payload.get("user_id")
        try:
            from bson.objectid import ObjectId
            user = await db.users.find_one({"_id": ObjectId(user_id)})
        except:
            user = await db.users.find_one({"id": user_id})
        
        if not user or not user.get("totp_enabled") or not user.get("totp_secret"):
            raise HTTPException(status_code=400, detail="2FA ist für diesen Account nicht aktiv")
            
        user["id"] = str(user["_id"])
        
        # Verify TOTP code
        totp = pyotp.TOTP(user["totp_secret"])
        if not totp.verify(data.totp_code):
            await log_audit("LOGIN_FAILED_2FA", user["id"], user.get("name", ""), {
                "ip_address": client_ip
            })
            raise HTTPException(status_code=401, detail="Falscher Authenticator-Code")
            
        # Success!
        await log_audit("USER_LOGIN", user["id"], user.get("name", user["email"]), {
            "email": user["email"],
            "role": user["role"],
            "ip_address": client_ip,
            "2fa_used": True
        })
        
        token = create_token(user["id"], user["role"])
        user.pop("password", None)
        user.pop("_id", None)
        user.pop("totp_secret", None)
        
        return TokenResponse(access_token=token, user=UserResponse(**user))
        
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Login-Sitzung abgelaufen, bitte erneut anmelden.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Ungültige Sitzung.")

# ==================== 2FA SETUP ====================

@api_router.post("/auth/2fa/setup")
async def setup_2fa(user: dict = Depends(get_current_user)):
    # Generate new secret
    secret = pyotp.random_base32()
    
    # Temporarily store it, but don't enable yet
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"temp_totp_secret": secret}}
    )
    
    # Generate provisioning URI
    totp = pyotp.TOTP(secret)
    provisioning_uri = totp.provisioning_uri(
        name=user["email"], 
        issuer_name="AbschleppApp"
    )
    
    # Create QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(provisioning_uri)
    qr.make(fit=True)
    
    # Convert QR Code to base64 for frontend
    img_buffer = BytesIO()
    img = qr.make_image(image_factory=PyPNGImage)
    img.save(img_buffer)
    img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
    
    return {
        "secret": secret,
        "qr_code": f"data:image/png;base64,{img_base64}"
    }

@api_router.post("/auth/2fa/verify-setup")
async def verify_2fa_setup(data: TOTPVerifyRequest, user: dict = Depends(get_current_user)):
    user_data = await db.users.find_one({"id": user["id"]})
    secret = user_data.get("temp_totp_secret")
    
    if not secret:
        raise HTTPException(status_code=400, detail="Kein 2FA-Setup im Gange")
        
    totp = pyotp.TOTP(secret)
    if not totp.verify(data.totp_code):
        raise HTTPException(status_code=400, detail="Falscher Code")
        
    # Successfully verified, make it permanent
    await db.users.update_one(
        {"id": user["id"]},
        {
            "$set": {
                "totp_secret": secret,
                "totp_enabled": True
            },
            "$unset": {"temp_totp_secret": ""}
        }
    )
    
    await log_audit("2FA_ENABLED", user["id"], user.get("name", ""), {})
    
    return {"message": "2FA erfolgreich aktiviert"}

@api_router.post("/auth/2fa/disable")
async def disable_2fa(data: TOTPVerifyRequest, user: dict = Depends(get_current_user)):
    user_data = await db.users.find_one({"id": user["id"]})
    
    if not user_data.get("totp_enabled"):
        raise HTTPException(status_code=400, detail="2FA ist nicht aktiviert")
        
    totp = pyotp.TOTP(user_data["totp_secret"])
    if not totp.verify(data.totp_code):
        raise HTTPException(status_code=400, detail="Falscher Code")
        
    await db.users.update_one(
        {"id": user["id"]},
        {
            "$set": {"totp_enabled": False},
            "$unset": {"totp_secret": ""}
        }
    )
    
    await log_audit("2FA_DISABLED", user["id"], user.get("name", ""), {})
    
    return {"message": "2FA wurde deaktiviert"}


# ==================== PASSWORD RESET ====================

@api_router.post("/auth/forgot-password")
async def forgot_password(data: PasswordResetRequest, background_tasks: BackgroundTasks):
    """Request password reset - sends email with reset link"""
    user = await db.users.find_one({"email": data.email})
    
    # Always return success to prevent email enumeration
    if not user:
        return {"message": "Falls ein Konto mit dieser E-Mail existiert, erhalten Sie einen Link zum Zurücksetzen."}
    
    # Generate reset token
    reset_token = generate_reset_token()
    expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
    
    # Store reset token in database
    await db.password_resets.delete_many({"user_id": user["id"]})  # Remove old tokens
    await db.password_resets.insert_one({
        "user_id": user["id"],
        "email": user["email"],
        "token": reset_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Send password reset email via BackgroundTasks
    try:
        background_tasks.add_task(
            email_service.send_password_reset_email,
            to_email=user["email"],
            reset_token=reset_token,
            user_name=user.get("name", user["email"])
        )
        logger.info(f"Password reset email queued for {data.email}")
    except Exception as e:
        logger.error(f"Failed to send password reset email: {e}")
        # Still return success to prevent email enumeration
    
    return {"message": "Falls ein Konto mit dieser E-Mail existiert, erhalten Sie einen Link zum Zurücksetzen."}

@api_router.post("/auth/reset-password")
async def reset_password(data: PasswordResetConfirm):
    """Reset password using token from email"""
    # Find valid reset token
    reset_record = await db.password_resets.find_one({"token": data.token})
    
    if not reset_record:
        raise HTTPException(status_code=400, detail="Ungültiger oder abgelaufener Link")
    
    # Check if token expired
    expires_at = datetime.fromisoformat(reset_record["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        await db.password_resets.delete_one({"token": data.token})
        raise HTTPException(status_code=400, detail="Der Link ist abgelaufen. Bitte fordern Sie einen neuen an.")
    
    # Validate new password
    is_valid, error_msg = validate_password(data.new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Update password
    new_hashed = hash_password(data.new_password)
    await db.users.update_one(
        {"id": reset_record["user_id"]},
        {"$set": {"password": new_hashed}}
    )
    
    # Audit log password reset
    await log_audit("PASSWORD_RESET", reset_record["user_id"], reset_record["email"], {
        "email": reset_record["email"],
        "method": "reset_token"
    })
    
    # Delete used token
    await db.password_resets.delete_one({"token": data.token})
    
    return {"message": "Passwort erfolgreich geändert. Sie können sich jetzt anmelden."}

@api_router.get("/auth/verify-reset-token/{token}")
async def verify_reset_token(token: str):
    """Verify if a reset token is valid"""
    reset_record = await db.password_resets.find_one({"token": token})
    
    if not reset_record:
        raise HTTPException(status_code=400, detail="Ungültiger Link")
    
    expires_at = datetime.fromisoformat(reset_record["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Der Link ist abgelaufen")
    
    return {"valid": True, "email": reset_record["email"]}

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(**user)

# ==================== TOWING SERVICE ROUTES ====================

@api_router.get("/services", response_model=List[UserResponse])
async def get_towing_services(user: dict = Depends(get_current_user)):
    if user["role"] == UserRole.AUTHORITY:
        # For employees, get linked services from the main authority account
        if user.get("is_main_authority"):
            linked_ids = user.get("linked_services", [])
        else:
            # Employee - get linked services from parent authority
            parent_id = user.get("parent_authority_id")
            if parent_id:
                # Try to find by UUID id first, then by MongoDB _id
                parent = await db.users.find_one({"id": parent_id})
                if not parent:
                    # Try with ObjectId
                    try:
                        from bson.objectid import ObjectId
                        parent = await db.users.find_one({"_id": ObjectId(parent_id)})
                    except:
                        parent = None
                linked_ids = parent.get("linked_services", []) if parent else []
            else:
                linked_ids = []
        
        if not linked_ids:
            return []
        services = await db.users.find(
            {"id": {"$in": linked_ids}, "role": UserRole.TOWING_SERVICE, "approval_status": ApprovalStatus.APPROVED}
        ).to_list(100)
        # Ensure id field exists for each service
        result = []
        for s in services:
            if "id" not in s and "_id" in s:
                s["id"] = str(s.pop("_id"))
            elif "_id" in s:
                s.pop("_id")
            if "password" in s:
                s.pop("password")
            result.append(UserResponse(**s))
        return result
    else:
        # Admin sees all approved services
        services = await db.users.find(
            {"role": UserRole.TOWING_SERVICE, "approval_status": ApprovalStatus.APPROVED},
            {"_id": 0, "password": 0}
        ).to_list(100)
    return [UserResponse(**s) for s in services]

@api_router.post("/services/link")
async def link_service(data: LinkServiceRequest, user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.AUTHORITY:
        raise HTTPException(status_code=403, detail="Only authorities can link services")
    
    # Only main authority can link services
    if not user.get("is_main_authority"):
        raise HTTPException(status_code=403, detail="Nur der Haupt-Account kann Abschleppdienste verknüpfen")
    
    # Only allow linking approved services
    service = await db.users.find_one({
        "service_code": data.service_code, 
        "role": UserRole.TOWING_SERVICE,
        "approval_status": ApprovalStatus.APPROVED
    })
    if not service:
        raise HTTPException(status_code=404, detail="Kein freigeschalteter Abschleppdienst mit diesem Code gefunden")
    
    linked = user.get("linked_services", [])
    if service["id"] in linked:
        raise HTTPException(status_code=400, detail="Service already linked")
    
    # Add service to authority's linked_services
    await db.users.update_one(
        {"id": user["id"]},
        {"$push": {"linked_services": service["id"]}}
    )
    
    # NEW: Also add authority to service's linked_authorities (bidirectional link)
    await db.users.update_one(
        {"id": service["id"]},
        {"$addToSet": {"linked_authorities": user["id"]}}
    )
    
    return {"message": "Service linked successfully", "service_name": service["company_name"]}

@api_router.delete("/services/unlink/{service_id}")
async def unlink_service(service_id: str, user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.AUTHORITY:
        raise HTTPException(status_code=403, detail="Only authorities can unlink services")
    
    # Only main authority can unlink services
    if not user.get("is_main_authority"):
        raise HTTPException(status_code=403, detail="Nur der Haupt-Account kann Abschleppdienste entfernen")
    
    # Remove service from authority's linked_services
    await db.users.update_one(
        {"id": user["id"]},
        {"$pull": {"linked_services": service_id}}
    )
    
    # NEW: Also remove authority from service's linked_authorities
    await db.users.update_one(
        {"id": service_id},
        {"$pull": {"linked_authorities": user["id"]}}
    )
    
    return {"message": "Service unlinked successfully"}

# NEW: Get linked authorities for towing service
@api_router.get("/towing/linked-authorities")
async def get_linked_authorities(user: dict = Depends(get_current_user)):
    """Get all authorities that have linked this towing service"""
    if user["role"] != UserRole.TOWING_SERVICE:
        raise HTTPException(status_code=403, detail="Nur Abschleppdienste können ihre verknüpften Behörden abrufen")
    
    linked_authority_ids = user.get("linked_authorities", [])
    if not linked_authority_ids:
        return []
    
    # Fetch authority details
    authorities = await db.users.find(
        {
            "id": {"$in": linked_authority_ids}, 
            "role": UserRole.AUTHORITY,
            "is_main_authority": True,
            "approval_status": ApprovalStatus.APPROVED
        },
        {"_id": 0, "password": 0, "business_license": 0}
    ).to_list(length=100)
    
    return [UserResponse(**a) for a in authorities]

# NEW: Update costs endpoint for towing service
@api_router.patch("/services/costs")
async def update_costs(data: UpdateCostsRequest, user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.TOWING_SERVICE:
        raise HTTPException(status_code=403, detail="Only towing services can update costs")
    
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"tow_cost": data.tow_cost, "daily_cost": data.daily_cost}}
    )
    
    # Fetch updated user
    updated_user = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password": 0})
    return UserResponse(**updated_user)

# NEW: Extended pricing settings endpoint
@api_router.patch("/services/pricing-settings")
async def update_pricing_settings(data: PricingSettingsRequest, user: dict = Depends(get_current_user)):
    """Update extended pricing settings for towing service"""
    if user["role"] != UserRole.TOWING_SERVICE:
        raise HTTPException(status_code=403, detail="Only towing services can update pricing")
    
    update_data = {}
    if data.time_based_enabled is not None:
        update_data["time_based_enabled"] = data.time_based_enabled
    if data.first_half_hour is not None:
        update_data["first_half_hour"] = data.first_half_hour
    if data.additional_half_hour is not None:
        update_data["additional_half_hour"] = data.additional_half_hour
    if data.tow_cost is not None:
        update_data["tow_cost"] = data.tow_cost
    if data.daily_cost is not None:
        update_data["daily_cost"] = data.daily_cost
    if data.processing_fee is not None:
        update_data["processing_fee"] = data.processing_fee
    if data.empty_trip_fee is not None:
        update_data["empty_trip_fee"] = data.empty_trip_fee
    if data.night_surcharge is not None:
        update_data["night_surcharge"] = data.night_surcharge
    if data.weekend_surcharge is not None:
        update_data["weekend_surcharge"] = data.weekend_surcharge
    if data.heavy_vehicle_surcharge is not None:
        update_data["heavy_vehicle_surcharge"] = data.heavy_vehicle_surcharge
    
    # NEW: Handle weight categories
    if data.weight_categories is not None:
        # Convert to dict list and generate IDs if missing
        weight_cats = []
        for cat in data.weight_categories:
            cat_dict = cat.model_dump() if hasattr(cat, 'model_dump') else cat.dict()
            if not cat_dict.get("id"):
                cat_dict["id"] = str(uuid.uuid4())
            weight_cats.append(cat_dict)
        update_data["weight_categories"] = weight_cats
    
    if update_data:
        await db.users.update_one({"id": user["id"]}, {"$set": update_data})
    
    updated_user = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password": 0})
    
    # Audit log
    await log_audit("PRICING_UPDATED", user["id"], user.get("company_name", user["name"]), {
        "settings": update_data
    })
    
    return UserResponse(**updated_user)

# NEW: Update company info (address, phone, email, opening hours)
class CompanyInfoUpdate(BaseModel):
    company_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    yard_address: Optional[str] = None
    yard_lat: Optional[float] = None
    yard_lng: Optional[float] = None
    opening_hours: Optional[str] = None

@api_router.patch("/towing/company-info")
async def update_company_info(data: CompanyInfoUpdate, user: dict = Depends(get_current_user)):
    """Update towing service company information"""
    if user["role"] != UserRole.TOWING_SERVICE:
        raise HTTPException(status_code=403, detail="Nur Abschleppdienste können Firmendaten ändern")
    
    update_data = {}
    if data.company_name is not None:
        update_data["company_name"] = data.company_name
    if data.phone is not None:
        update_data["phone"] = data.phone
    if data.email is not None:
        update_data["email"] = data.email
    if data.yard_address is not None:
        update_data["yard_address"] = data.yard_address
    if data.yard_lat is not None:
        update_data["yard_lat"] = data.yard_lat
    if data.yard_lng is not None:
        update_data["yard_lng"] = data.yard_lng
    if data.opening_hours is not None:
        update_data["opening_hours"] = data.opening_hours
    
    if update_data:
        await db.users.update_one({"id": user["id"]}, {"$set": update_data})
    
    updated_user = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password": 0})
    
    # Audit log
    await log_audit("COMPANY_INFO_UPDATED", user["id"], user.get("company_name", user["name"]), {
        "updated_fields": list(update_data.keys())
    })
    
    return UserResponse(**updated_user)

# ==================== FAHRZEUGKATEGORIEN-PREISE ====================

@api_router.get("/vehicle-categories")
async def get_vehicle_categories(user: dict = Depends(get_current_user)):
    """Hole alle Fahrzeugkategorien mit Preisen für den aktuellen Benutzer"""
    if user["role"] not in [UserRole.AUTHORITY, UserRole.TOWING_SERVICE]:
        raise HTTPException(status_code=403, detail="Nicht berechtigt")
    
    # Hole Kategorien für diesen Benutzer
    categories = await db.vehicle_categories.find({"owner_id": user["id"]}).to_list(100)
    
    # Entferne MongoDB _id
    for cat in categories:
        cat.pop("_id", None)
    
    return categories

@api_router.post("/vehicle-categories")
async def create_vehicle_category(category: VehicleCategoryPrice, user: dict = Depends(get_current_user)):
    """Neue Fahrzeugkategorie mit Preisen anlegen"""
    if user["role"] not in [UserRole.AUTHORITY, UserRole.TOWING_SERVICE]:
        raise HTTPException(status_code=403, detail="Nicht berechtigt")
    
    new_category = {
        "id": str(uuid.uuid4()),
        "owner_id": user["id"],
        "owner_type": user["role"],
        "name": category.name,
        "description": category.description,
        "base_price": category.base_price,
        "daily_rate": category.daily_rate,
        "is_active": category.is_active,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.vehicle_categories.insert_one(new_category)
    
    await log_audit("VEHICLE_CATEGORY_CREATED", user["id"], user.get("name", ""), {
        "category_name": category.name,
        "base_price": category.base_price,
        "daily_rate": category.daily_rate
    })
    
    new_category.pop("_id", None)
    return new_category

@api_router.put("/vehicle-categories/{category_id}")
async def update_vehicle_category(category_id: str, category: VehicleCategoryPrice, user: dict = Depends(get_current_user)):
    """Fahrzeugkategorie aktualisieren"""
    if user["role"] not in [UserRole.AUTHORITY, UserRole.TOWING_SERVICE]:
        raise HTTPException(status_code=403, detail="Nicht berechtigt")
    
    existing = await db.vehicle_categories.find_one({"id": category_id, "owner_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Kategorie nicht gefunden")
    
    update_data = {
        "name": category.name,
        "description": category.description,
        "base_price": category.base_price,
        "daily_rate": category.daily_rate,
        "is_active": category.is_active,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.vehicle_categories.update_one({"id": category_id}, {"$set": update_data})
    
    await log_audit("VEHICLE_CATEGORY_UPDATED", user["id"], user.get("name", ""), {
        "category_id": category_id,
        "category_name": category.name
    })
    
    updated = await db.vehicle_categories.find_one({"id": category_id}, {"_id": 0})
    return updated

@api_router.delete("/vehicle-categories/{category_id}")
async def delete_vehicle_category(category_id: str, user: dict = Depends(get_current_user)):
    """Fahrzeugkategorie löschen"""
    if user["role"] not in [UserRole.AUTHORITY, UserRole.TOWING_SERVICE]:
        raise HTTPException(status_code=403, detail="Nicht berechtigt")
    
    existing = await db.vehicle_categories.find_one({"id": category_id, "owner_id": user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Kategorie nicht gefunden")
    
    await db.vehicle_categories.delete_one({"id": category_id})
    
    await log_audit("VEHICLE_CATEGORY_DELETED", user["id"], user.get("name", ""), {
        "category_id": category_id,
        "category_name": existing.get("name")
    })
    
    return {"status": "deleted"}

@api_router.post("/vehicle-categories/calculate")
async def calculate_category_price(
    category_id: str = None,
    days: int = 1,
    user: dict = Depends(get_current_user)
):
    """Berechne Preis basierend auf Kategorie und Standtagen"""
    if not category_id:
        raise HTTPException(status_code=400, detail="category_id erforderlich")
    
    category = await db.vehicle_categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Kategorie nicht gefunden")
    
    base_price = category.get("base_price", 0)
    daily_rate = category.get("daily_rate", 0)
    
    # Erste 24h = Grundpreis, danach Tagessatz
    if days <= 1:
        total = base_price
    else:
        total = base_price + ((days - 1) * daily_rate)
    
    return {
        "category_id": category_id,
        "category_name": category.get("name"),
        "days": days,
        "base_price": base_price,
        "daily_rate": daily_rate,
        "total": total,
        "breakdown": [
            {"label": f"Grundpreis (erste 24h)", "amount": base_price},
            {"label": f"{days - 1} × weitere 24h à {daily_rate}€", "amount": (days - 1) * daily_rate} if days > 1 else None
        ]
    }

# NEW: Get weight categories for a specific towing service (for authorities to use in job creation)
@api_router.get("/services/{service_id}/weight-categories")
async def get_service_weight_categories(service_id: str, user: dict = Depends(get_current_user)):
    """Get weight categories defined by a towing service for use in job creation dropdown"""
    if user["role"] not in [UserRole.AUTHORITY, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Nicht berechtigt")
    
    service = await db.users.find_one({"id": service_id, "role": UserRole.TOWING_SERVICE})
    if not service:
        raise HTTPException(status_code=404, detail="Abschleppdienst nicht gefunden")
    
    weight_categories = service.get("weight_categories", [])
    return {"weight_categories": weight_categories, "service_name": service.get("company_name")}

# NEW: Calculate job costs based on service pricing OR vehicle category
@api_router.get("/jobs/{job_id}/calculate-costs")
async def calculate_job_costs(job_id: str, user: dict = Depends(get_current_user)):
    """Calculate costs for a job based on vehicle category pricing or towing service's pricing settings"""
    job = await db.jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Auftrag nicht gefunden")
    
    breakdown = []
    total = 0.0
    
    # NEW: Check if job has a vehicle category assigned (new dynamic pricing)
    vehicle_category_id = job.get("vehicle_category_id")
    if vehicle_category_id:
        category = await db.vehicle_categories.find_one({"id": vehicle_category_id})
        if category:
            base_price = category.get("base_price", 0)
            daily_rate = category.get("daily_rate", 0)
            category_name = category.get("name", "Unbekannt")
            
            # Calculate days in yard
            days = 1
            if job.get("in_yard_at"):
                in_yard_date = datetime.fromisoformat(job["in_yard_at"].replace("Z", "+00:00"))
                now = datetime.now(timezone.utc)
                days = max(1, (now - in_yard_date).days + 1)
            elif job.get("towed_at"):
                # Fallback: use towed_at if in_yard_at not set
                towed_date = datetime.fromisoformat(job["towed_at"].replace("Z", "+00:00"))
                now = datetime.now(timezone.utc)
                days = max(1, (now - towed_date).days + 1)
            
            # Add base price (first 24h)
            breakdown.append({"label": f"{category_name} - Grundpreis (erste 24h)", "amount": base_price})
            total += base_price
            
            # Add daily rate for additional days
            if days > 1 and daily_rate > 0:
                additional_days_cost = (days - 1) * daily_rate
                breakdown.append({"label": f"{days - 1} × weitere 24h à {daily_rate:.2f}€", "amount": additional_days_cost})
                total += additional_days_cost
            
            # Note: We can still add service-specific surcharges on top
            service = await db.users.find_one({"id": job.get("assigned_service_id")})
            if service:
                # Night surcharge (check if towed at night: 22:00 - 06:00)
                if job.get("towed_at"):
                    towed_time = datetime.fromisoformat(job["towed_at"].replace("Z", "+00:00"))
                    hour = towed_time.hour
                    if hour >= 22 or hour < 6:
                        night_surcharge = service.get("night_surcharge", 0) or 0
                        if night_surcharge > 0:
                            breakdown.append({"label": "Nachtzuschlag", "amount": night_surcharge})
                            total += night_surcharge
                
                # Weekend surcharge (Saturday/Sunday)
                if job.get("towed_at"):
                    towed_time = datetime.fromisoformat(job["towed_at"].replace("Z", "+00:00"))
                    if towed_time.weekday() >= 5:  # Saturday = 5, Sunday = 6
                        weekend_surcharge = service.get("weekend_surcharge", 0) or 0
                        if weekend_surcharge > 0:
                            breakdown.append({"label": "Wochenendzuschlag", "amount": weekend_surcharge})
                            total += weekend_surcharge
                
                # Processing fee
                processing_fee = service.get("processing_fee", 0) or 0
                if processing_fee > 0:
                    breakdown.append({"label": "Bearbeitungsgebühr", "amount": processing_fee})
                    total += processing_fee
            
            return {"total": round(total, 2), "breakdown": breakdown, "pricing_source": "vehicle_category", "category_name": category_name}
    
    # FALLBACK: Original service-based pricing logic
    service = await db.users.find_one({"id": job.get("assigned_service_id")})
    if not service:
        return {"total": 0, "breakdown": [], "pricing_source": "none"}
    
    # Check if time-based pricing is enabled
    time_based_applied = False
    if service.get("time_based_enabled") and job.get("accepted_at") and job.get("in_yard_at"):
        accepted = datetime.fromisoformat(job["accepted_at"].replace("Z", "+00:00"))
        in_yard = datetime.fromisoformat(job["in_yard_at"].replace("Z", "+00:00"))
        duration_minutes = (in_yard - accepted).total_seconds() / 60
        half_hours = max(1, int((duration_minutes + 29) / 30))  # Round up
        
        first_hh = service.get("first_half_hour", 0) or 0
        add_hh = service.get("additional_half_hour", 0) or 0
        
        if first_hh > 0:
            breakdown.append({"label": "Erste halbe Stunde", "amount": first_hh})
            total += first_hh
            time_based_applied = True
            
            if half_hours > 1 and add_hh > 0:
                additional = (half_hours - 1) * add_hh
                breakdown.append({"label": f"{half_hours - 1} × weitere halbe Stunde", "amount": additional})
                total += additional
                
    if not time_based_applied:
        # Standard pricing fallback
        tow_cost = service.get("tow_cost", 0) or 0
        if tow_cost > 0:
            breakdown.append({"label": "Anfahrt/Abschleppen", "amount": tow_cost})
            total += tow_cost
    
    # Daily costs (if in yard)
    if job.get("in_yard_at"):
        daily_cost = service.get("daily_cost", 0) or 0
        if daily_cost > 0:
            in_yard_date = datetime.fromisoformat(job["in_yard_at"].replace("Z", "+00:00"))
            now = datetime.now(timezone.utc)
            days = max(1, (now - in_yard_date).days + 1)
            daily_total = days * daily_cost
            breakdown.append({"label": f"Standkosten ({days} Tag{'e' if days > 1 else ''})", "amount": daily_total})
            total += daily_total
    
    # Processing fee
    processing_fee = service.get("processing_fee", 0) or 0
    if processing_fee > 0:
        breakdown.append({"label": "Bearbeitungsgebühr", "amount": processing_fee})
        total += processing_fee
    
    # Empty trip fee
    if job.get("is_empty_trip"):
        empty_fee = service.get("empty_trip_fee", 0) or 0
        if empty_fee > 0:
            breakdown.append({"label": "Leerfahrt", "amount": empty_fee})
            total += empty_fee
    
    # Heavy vehicle surcharge - NEW: Use weight_category if available, otherwise fall back to old logic
    weight_category_surcharge_applied = False
    if job.get("weight_category_surcharge") and job.get("weight_category_surcharge") > 0:
        weight_cat_name = job.get("weight_category_name", "Gewichtskategorie")
        weight_surcharge = job.get("weight_category_surcharge", 0)
        breakdown.append({"label": f"Zuschlag {weight_cat_name}", "amount": weight_surcharge})
        total += weight_surcharge
        weight_category_surcharge_applied = True
    
    # Fallback to old heavy_vehicle_surcharge if no weight category was used
    if not weight_category_surcharge_applied and job.get("vehicle_category") == "over_3_5t":
        heavy_surcharge = service.get("heavy_vehicle_surcharge", 0) or 0
        if heavy_surcharge > 0:
            breakdown.append({"label": "Schwerlastzuschlag (ab 3,5t)", "amount": heavy_surcharge})
            total += heavy_surcharge
    
    # Night surcharge (check if towed at night: 22:00 - 06:00)
    if job.get("towed_at"):
        towed_time = datetime.fromisoformat(job["towed_at"].replace("Z", "+00:00"))
        hour = towed_time.hour
        if hour >= 22 or hour < 6:
            night_surcharge = service.get("night_surcharge", 0) or 0
            if night_surcharge > 0:
                breakdown.append({"label": "Nachtzuschlag", "amount": night_surcharge})
                total += night_surcharge
    
    # Weekend surcharge (Saturday/Sunday)
    if job.get("towed_at"):
        towed_time = datetime.fromisoformat(job["towed_at"].replace("Z", "+00:00"))
        if towed_time.weekday() >= 5:  # Saturday = 5, Sunday = 6
            weekend_surcharge = service.get("weekend_surcharge", 0) or 0
            if weekend_surcharge > 0:
                breakdown.append({"label": "Wochenendzuschlag", "amount": weekend_surcharge})
                total += weekend_surcharge
    
    return {"total": round(total, 2), "breakdown": breakdown, "pricing_source": "service"}

# ==================== AUTHORITY SETTINGS & PRICING ====================

class AuthorityPriceCategory(BaseModel):
    id: Optional[str] = None
    name: str  # z.B. "PKW bis 4t", "LKW 4-7,5t"
    base_price: float  # Grundpreis (erste 24h)
    daily_rate: float  # Preis pro weitere 24h
    is_active: bool = True

# NEW: Authority yard location
class AuthorityYard(BaseModel):
    id: Optional[str] = None
    name: str  # z.B. "Haupthof", "Außenstelle Nord"
    address: str
    lat: Optional[float] = None
    lng: Optional[float] = None
    phone: Optional[str] = None
    is_active: bool = True

class AuthoritySettingsUpdate(BaseModel):
    yard_model: Optional[str] = None  # "authority_yard" or "service_yard"
    price_categories: Optional[List[AuthorityPriceCategory]] = None
    yards: Optional[List[AuthorityYard]] = None  # Multiple yards
    # Legacy single yard (deprecated, use yards instead)
    yard_address: Optional[str] = None
    yard_lat: Optional[float] = None
    yard_lng: Optional[float] = None

@api_router.patch("/authority/settings")
async def update_authority_settings(data: AuthoritySettingsUpdate, user: dict = Depends(get_current_user)):
    """Update authority settings (yard model, pricing, etc.)"""
    if user["role"] != UserRole.AUTHORITY:
        raise HTTPException(status_code=403, detail="Nur Behörden können Einstellungen ändern")
    
    # Get the main authority (either self or parent)
    authority_id = get_authority_id(user)
    
    update_data = {}
    
    if data.yard_model is not None:
        if data.yard_model not in ["authority_yard", "service_yard"]:
            raise HTTPException(status_code=400, detail="Ungültiges Hof-Modell")
        update_data["yard_model"] = data.yard_model
    
    if data.price_categories is not None:
        # Convert to dict list and generate IDs if missing
        price_cats = []
        for cat in data.price_categories:
            cat_dict = cat.model_dump() if hasattr(cat, 'model_dump') else cat.dict()
            if not cat_dict.get("id"):
                cat_dict["id"] = str(uuid.uuid4())
            price_cats.append(cat_dict)
        update_data["price_categories"] = price_cats
    
    # NEW: Handle multiple yards
    if data.yards is not None:
        yards_list = []
        for yard in data.yards:
            yard_dict = yard.model_dump() if hasattr(yard, 'model_dump') else yard.dict()
            if not yard_dict.get("id"):
                yard_dict["id"] = str(uuid.uuid4())
            yards_list.append(yard_dict)
        update_data["yards"] = yards_list
    
    # Legacy single yard support
    if data.yard_address is not None:
        update_data["yard_address"] = data.yard_address
    if data.yard_lat is not None:
        update_data["yard_lat"] = data.yard_lat
    if data.yard_lng is not None:
        update_data["yard_lng"] = data.yard_lng
    
    if update_data:
        await db.users.update_one({"id": authority_id}, {"$set": update_data})
        # Also update employees to inherit settings
        if data.yard_model is not None:
            await db.users.update_many(
                {"parent_authority_id": authority_id},
                {"$set": {"yard_model": data.yard_model}}
            )
    
    updated_user = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password": 0})
    
    await log_audit("AUTHORITY_SETTINGS_UPDATED", user["id"], user["name"], {
        "settings": list(update_data.keys())
    })
    
    return UserResponse(**updated_user)

@api_router.get("/authority/settings")
async def get_authority_settings(user: dict = Depends(get_current_user)):
    """Get authority settings including pricing categories"""
    if user["role"] != UserRole.AUTHORITY:
        raise HTTPException(status_code=403, detail="Nur Behörden haben Zugriff")
    
    authority_id = get_authority_id(user)
    authority = await db.users.find_one({"id": authority_id}, {"_id": 0, "password": 0})
    
    return {
        "yard_model": authority.get("yard_model", "service_yard"),  # Default: Abschleppdienst-Hof
        "price_categories": authority.get("price_categories", []),
        "yards": authority.get("yards", []),  # Multiple yards
        "yard_address": authority.get("yard_address"),
        "yard_lat": authority.get("yard_lat"),
        "yard_lng": authority.get("yard_lng")
    }

# ==================== AUTHORITY EMPLOYEE MANAGEMENT ====================

def get_authority_id(user: dict) -> str:
    """Get the main authority ID for a user (either main authority or employee)"""
    if user.get("is_main_authority"):
        return user["id"]
    return user.get("parent_authority_id", user["id"])

@api_router.post("/authority/employees", response_model=EmployeeResponse)
async def create_employee(data: CreateEmployeeRequest, user: dict = Depends(get_current_user)):
    """Create a new employee for the authority"""
    if user["role"] != UserRole.AUTHORITY:
        raise HTTPException(status_code=403, detail="Nur Behörden können Mitarbeiter erstellen")
    
    # Only main authority can create employees
    if not user.get("is_main_authority"):
        raise HTTPException(status_code=403, detail="Nur der Haupt-Account kann Mitarbeiter erstellen")
    
    # Check if email exists
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="E-Mail bereits registriert")
    
    employee_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Generate Dienstnummer
    dienstnummer = await generate_dienstnummer(user["id"])
    
    employee_doc = {
        "id": employee_id,
        "email": data.email,
        "password": hash_password(data.password),
        "role": UserRole.AUTHORITY,
        "name": data.name,
        "created_at": now,
        "authority_name": user.get("authority_name"),
        "department": user.get("department"),
        "linked_services": user.get("linked_services", []),  # Inherit linked services
        "is_main_authority": False,
        "parent_authority_id": user["id"],
        "dienstnummer": dienstnummer
    }
    
    await db.users.insert_one(employee_doc)
    
    # Audit log employee creation
    await log_audit("EMPLOYEE_CREATED", user["id"], user["name"], {
        "employee_id": employee_id,
        "employee_email": data.email,
        "employee_name": data.name,
        "dienstnummer": dienstnummer,
        "authority_id": user["id"]
    })
    
    return EmployeeResponse(
        id=employee_id,
        email=data.email,
        name=data.name,
        dienstnummer=dienstnummer,
        is_blocked=False,
        created_at=now
    )

@api_router.get("/authority/employees", response_model=List[EmployeeResponse])
async def get_employees(user: dict = Depends(get_current_user)):
    """Get all employees for the authority"""
    if user["role"] != UserRole.AUTHORITY:
        raise HTTPException(status_code=403, detail="Nur Behörden")
    
    # Only main authority can view employees
    if not user.get("is_main_authority"):
        raise HTTPException(status_code=403, detail="Nur der Haupt-Account kann Mitarbeiter sehen")
    
    employees = await db.users.find(
        {"parent_authority_id": user["id"]},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    return [EmployeeResponse(
        id=e["id"],
        email=e["email"],
        name=e["name"],
        dienstnummer=e.get("dienstnummer", ""),
        is_blocked=e.get("is_blocked", False),
        created_at=e["created_at"]
    ) for e in employees]

@api_router.patch("/authority/employees/{employee_id}/block")
async def block_employee(employee_id: str, data: AdminBlockUserRequest, user: dict = Depends(get_current_user)):
    """Block/unblock an employee"""
    if user["role"] != UserRole.AUTHORITY or not user.get("is_main_authority"):
        raise HTTPException(status_code=403, detail="Nur der Haupt-Account kann Mitarbeiter sperren")
    
    employee = await db.users.find_one({"id": employee_id, "parent_authority_id": user["id"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Mitarbeiter nicht gefunden")
    
    await db.users.update_one({"id": employee_id}, {"$set": {"is_blocked": data.blocked}})
    
    action = "gesperrt" if data.blocked else "entsperrt"
    
    # Audit log employee block/unblock
    await log_audit("EMPLOYEE_BLOCKED" if data.blocked else "EMPLOYEE_UNBLOCKED", user["id"], user["name"], {
        "employee_id": employee_id,
        "employee_name": employee["name"],
        "blocked": data.blocked
    })
    
    return {"message": f"Mitarbeiter {employee['name']} wurde {action}"}

@api_router.delete("/authority/employees/{employee_id}")
async def delete_employee(employee_id: str, user: dict = Depends(get_current_user)):
    """Delete an employee"""
    if user["role"] != UserRole.AUTHORITY or not user.get("is_main_authority"):
        raise HTTPException(status_code=403, detail="Nur der Haupt-Account kann Mitarbeiter löschen")
    
    employee = await db.users.find_one({"id": employee_id, "parent_authority_id": user["id"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Mitarbeiter nicht gefunden")
    
    await db.users.delete_one({"id": employee_id})
    
    # Audit log employee deletion
    await log_audit("EMPLOYEE_DELETED", user["id"], user["name"], {
        "employee_id": employee_id,
        "employee_name": employee["name"],
        "employee_email": employee.get("email")
    })
    
    return {"message": f"Mitarbeiter {employee['name']} wurde gelöscht"}

@api_router.patch("/authority/employees/{employee_id}/password")
async def update_employee_password(employee_id: str, data: AdminUpdatePasswordRequest, user: dict = Depends(get_current_user)):
    """Update an employee's password"""
    if user["role"] != UserRole.AUTHORITY or not user.get("is_main_authority"):
        raise HTTPException(status_code=403, detail="Nur der Haupt-Account kann Passwörter ändern")
    
    employee = await db.users.find_one({"id": employee_id, "parent_authority_id": user["id"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Mitarbeiter nicht gefunden")
    
    new_hashed = hash_password(data.new_password)
    await db.users.update_one({"id": employee_id}, {"$set": {"password": new_hashed}})
    
    # Audit log employee password change
    await log_audit("EMPLOYEE_PASSWORD_CHANGED", user["id"], user["name"], {
        "employee_id": employee_id,
        "employee_name": employee["name"]
    })
    
    return {"message": f"Passwort für {employee['name']} wurde aktualisiert"}

# ==================== ADMIN APPROVAL ROUTES ====================

@api_router.get("/admin/pending-services", response_model=List[UserResponse])
async def get_pending_services(user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    services = await db.users.find(
        {"role": UserRole.TOWING_SERVICE, "approval_status": ApprovalStatus.PENDING},
        {"_id": 0, "password": 0}
    ).to_list(100)
    return [UserResponse(**s) for s in services]

@api_router.get("/admin/pending-authorities", response_model=List[UserResponse])
async def get_pending_authorities(user: dict = Depends(get_current_user)):
    """Get all authorities waiting for approval"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    authorities = await db.users.find(
        {"role": UserRole.AUTHORITY, "approval_status": ApprovalStatus.PENDING, "is_main_authority": True},
        {"_id": 0, "password": 0}
    ).to_list(100)
    return [UserResponse(**a) for a in authorities]

# NEW: Sync authority-service links (ensures bidirectional linking)
@api_router.post("/admin/sync-links")
async def sync_authority_service_links(user: dict = Depends(get_current_user)):
    """Synchronize bidirectional links between authorities and towing services"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Find all authorities with linked services
    authorities = await db.users.find(
        {"role": UserRole.AUTHORITY, "linked_services": {"$exists": True, "$ne": []}},
        {"_id": 0}
    ).to_list(1000)
    
    updated_count = 0
    for auth in authorities:
        linked_services = auth.get("linked_services", [])
        for service_id in linked_services:
            # Add authority to service's linked_authorities
            result = await db.users.update_one(
                {"id": service_id, "role": UserRole.TOWING_SERVICE},
                {"$addToSet": {"linked_authorities": auth["id"]}}
            )
            if result.modified_count > 0:
                updated_count += 1
    
    await log_audit("LINKS_SYNCHRONIZED", user["id"], user["name"], {
        "updated_count": updated_count
    })
    
    return {"message": f"Synchronisierung abgeschlossen. {updated_count} Verknüpfungen aktualisiert."}

@api_router.post("/admin/approve-service/{service_id}")
async def approve_service(service_id: str, data: ApproveServiceRequest, background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    service = await db.users.find_one({"id": service_id, "role": UserRole.TOWING_SERVICE})
    if not service:
        raise HTTPException(status_code=404, detail="Towing service not found")
    
    if data.approved:
        await db.users.update_one(
            {"id": service_id},
            {"$set": {"approval_status": ApprovalStatus.APPROVED, "rejection_reason": None}}
        )
        # Audit log approval
        await log_audit("SERVICE_APPROVED", user["id"], user["name"], {
            "service_id": service_id,
            "company_name": service["company_name"]
        })
        
        # Send approval email via BackgroundTasks
        try:
            background_tasks.add_task(
                email_service.send_account_approved_email,
                to_email=service["email"],
                user_name=service.get("name", service["company_name"]),
                role=UserRole.TOWING_SERVICE
            )
            logger.info(f"Approval email queued for {service['email']}")
        except Exception as e:
            logger.error(f"Failed to send approval email: {e}")
        
        return {"message": f"{service['company_name']} wurde freigeschaltet"}
    else:
        await db.users.update_one(
            {"id": service_id},
            {"$set": {"approval_status": ApprovalStatus.REJECTED, "rejection_reason": data.rejection_reason}}
        )
        # Audit log rejection
        await log_audit("SERVICE_REJECTED", user["id"], user["name"], {
            "service_id": service_id,
            "company_name": service["company_name"],
            "reason": data.rejection_reason
        })
        return {"message": f"{service['company_name']} wurde abgelehnt"}

@api_router.post("/admin/approve-authority/{authority_id}")
async def approve_authority(authority_id: str, data: ApproveServiceRequest, background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    """Approve or reject an authority registration"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    authority = await db.users.find_one({"id": authority_id, "role": UserRole.AUTHORITY})
    if not authority:
        raise HTTPException(status_code=404, detail="Behörde nicht gefunden")
    
    if data.approved:
        await db.users.update_one(
            {"id": authority_id},
            {"$set": {"approval_status": ApprovalStatus.APPROVED, "rejection_reason": None}}
        )
        # Audit log approval
        await log_audit("AUTHORITY_APPROVED", user["id"], user["name"], {
            "authority_id": authority_id,
            "authority_name": authority["authority_name"]
        })
        
        # Send approval email via BackgroundTasks
        try:
            background_tasks.add_task(
                email_service.send_account_approved_email,
                to_email=authority["email"],
                user_name=authority.get("name", authority["authority_name"]),
                role=UserRole.AUTHORITY
            )
            logger.info(f"Authority approval email queued for {authority['email']}")
        except Exception as e:
            logger.error(f"Failed to send authority approval email: {e}")
        
        return {"message": f"{authority['authority_name']} wurde freigeschaltet"}
    else:
        await db.users.update_one(
            {"id": authority_id},
            {"$set": {"approval_status": ApprovalStatus.REJECTED, "rejection_reason": data.rejection_reason}}
        )
        # Audit log rejection
        await log_audit("AUTHORITY_REJECTED", user["id"], user["name"], {
            "authority_id": authority_id,
            "authority_name": authority["authority_name"],
            "reason": data.rejection_reason
        })
        return {"message": f"{authority['authority_name']} wurde abgelehnt"}

# ==================== JOB ROUTES ====================

@api_router.post("/jobs", response_model=JobResponse)
async def create_job(
    data: JobCreate, 
    user: dict = Depends(get_current_user),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key")
):
    # Allow authorities, admins, and towing services to create jobs
    if user["role"] not in [UserRole.AUTHORITY, UserRole.ADMIN, UserRole.TOWING_SERVICE]:
        raise HTTPException(status_code=403, detail="Keine Berechtigung zum Erstellen von Aufträgen")
    
    # ========== IDEMPOTENCY CHECK ==========
    if idempotency_key:
        existing_job = await db.jobs.find_one({"idempotency_key": idempotency_key})
        if existing_job:
            # If the exact same request was already processed, just return the existing result (200 OK)
            # This prevents duplicate inserts if the user clicked "Save" 5 times in panic.
            return JobResponse(**existing_job)
            
    # ========== DUPLICATE LICENSE PLATE CHECK ==========
    # Check if there's an active job with the same license plate
    # A job is considered "active" if it's not in 'released' status
    normalized_plate = data.license_plate.upper().replace(" ", "").replace("-", "")
    
    # Find any job with this license plate that is NOT released
    active_job = await db.jobs.find_one({
        "$or": [
            {"license_plate": data.license_plate.upper()},
            {"license_plate": normalized_plate},
            {"license_plate": {"$regex": f"^{re.escape(normalized_plate)}$", "$options": "i"}}
        ],
        "status": {"$ne": JobStatus.RELEASED}
    })
    
    if active_job:
        raise HTTPException(
            status_code=400, 
            detail=f"Ein Fahrzeug mit diesem Kennzeichen ({active_job['license_plate']}) ist bereits im System und wurde noch nicht freigegeben. Status: {active_job['status']}"
        )
    # ========== END DUPLICATE CHECK ==========
    
    job_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Handle towing service creating job for an authority
    if user["role"] == UserRole.TOWING_SERVICE:
        if not data.for_authority_id:
            raise HTTPException(status_code=400, detail="Abschleppdienste müssen eine Behörde angeben")
        
        # Verify the authority is linked to this towing service
        linked_authorities = user.get("linked_authorities", [])
        if data.for_authority_id not in linked_authorities:
            raise HTTPException(status_code=403, detail="Sie können nur Aufträge für verknüpfte Behörden erstellen")
        
        # Get authority details
        authority = await db.users.find_one({"id": data.for_authority_id, "role": UserRole.AUTHORITY})
        if not authority:
            raise HTTPException(status_code=404, detail="Behörde nicht gefunden")
        
        authority_id = data.for_authority_id
        authority_name = authority.get("authority_name")
        created_by_authority = authority_name
        created_by_dienstnummer = None  # Service doesn't have Dienstnummer
        # Job is auto-assigned to this towing service
        assigned_service_id = user["id"]
        assigned_service_name = user.get("company_name")
        # For service-created jobs, start in "assigned" status (already accepted)
        initial_status = JobStatus.ASSIGNED
    else:
        # Authority or Admin creating job
        # Get assigned service name if provided
        assigned_service_name = None
        assigned_service_id = data.assigned_service_id
        if assigned_service_id:
            service = await db.users.find_one({"id": assigned_service_id})
            if service:
                assigned_service_name = service.get("company_name")
        
        # Get the authority ID (either main or from parent)
        authority_id = get_authority_id(user)
        authority_name = user.get("authority_name")
        created_by_authority = authority_name
        created_by_dienstnummer = user.get("dienstnummer")
        initial_status = JobStatus.ASSIGNED if assigned_service_id else JobStatus.PENDING
    
    # Compress photos before storing
    compressed_photos = []
    for photo in data.photos:
        if photo and photo.startswith('data:image'):
            compressed_photos.append(compress_image_base64(photo))
        else:
            compressed_photos.append(photo)
    
    job_doc = {
        "id": job_id,
        "job_number": generate_job_number(),
        "license_plate": data.license_plate.upper(),
        "vin": data.vin.upper() if data.vin else None,
        "tow_reason": data.tow_reason,
        "location_address": data.location_address,
        "location_lat": data.location_lat,
        "location_lng": data.location_lng,
        "photos": compressed_photos,
        "notes": data.notes,
        "status": initial_status,
        "created_by_id": user["id"],
        "created_by_name": user["name"],
        "created_by_authority": created_by_authority,
        "created_by_dienstnummer": created_by_dienstnummer,
        "authority_id": authority_id,
        "assigned_service_id": assigned_service_id,
        "assigned_service_name": assigned_service_name,
        "service_notes": None,
        "service_photos": [],
        "owner_first_name": None,
        "owner_last_name": None,
        "owner_address": None,
        "payment_method": None,
        "payment_amount": None,
        "created_at": now,
        "updated_at": now,
        "on_site_at": None,
        "towed_at": None,
        "in_yard_at": None,
        "released_at": None,
        "accepted_at": now if user["role"] == UserRole.TOWING_SERVICE else None,  # Auto-accepted for service-created jobs
        # NEW: Job type and Sicherstellung fields
        "job_type": data.job_type or "towing",
        "sicherstellung_reason": data.sicherstellung_reason,
        "vehicle_category": data.vehicle_category,
        "ordering_authority": data.ordering_authority,
        "contact_attempts": data.contact_attempts,
        "contact_attempts_notes": data.contact_attempts_notes,
        "estimated_vehicle_value": data.estimated_vehicle_value,
        "is_empty_trip": False,
        "calculated_costs": None,
        # NEW: Vehicle category for dynamic pricing
        "vehicle_category_id": data.vehicle_category_id,
        # NEW: Target yard model
        "target_yard": data.target_yard or "service_yard",
        # Authority yard location (when target_yard = "authority_yard")
        "authority_yard_id": data.authority_yard_id,
        "authority_yard_name": data.authority_yard_name,
        "authority_yard_address": data.authority_yard_address,
        "authority_yard_lat": data.authority_yard_lat,
        "authority_yard_lng": data.authority_yard_lng,
        "authority_yard_phone": data.authority_yard_phone,
        # Authority pricing (when target_yard = "authority_yard")
        "authority_price_category_id": data.authority_price_category_id,
        "authority_price_category_name": data.authority_price_category_name,
        "authority_base_price": data.authority_base_price,
        "authority_daily_rate": data.authority_daily_rate,
        # NEW: Weight category for flexible surcharges (when target_yard = "service_yard")
        "weight_category_id": data.weight_category_id,
        "weight_category_name": data.weight_category_name,
        "weight_category_surcharge": data.weight_category_surcharge,
        # NEW: Track if job was created by towing service
        "created_by_service": user["role"] == UserRole.TOWING_SERVICE,
        "idempotency_key": idempotency_key
    }
    
    await db.jobs.insert_one(job_doc)
    job_doc.pop("_id", None)
    
    # Log audit
    await log_audit("JOB_CREATED", user["id"], user["name"], {
        "job_id": job_id,
        "job_number": job_doc["job_number"],
        "license_plate": job_doc["license_plate"]
    })
    
    return JobResponse(**job_doc)

@api_router.get("/jobs", response_model=List[JobResponse])
async def get_jobs(
    status: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    service_id: Optional[str] = None,  # NEW: Filter by towing service
    page: int = 1,
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    query = {}
    
    # Filter by role
    if user["role"] == UserRole.AUTHORITY:
        # Main authority sees all jobs from their authority, employees see only their own
        if user.get("is_main_authority"):
            query["authority_id"] = user["id"]
        else:
            query["created_by_id"] = user["id"]
    elif user["role"] == UserRole.TOWING_SERVICE:
        query["assigned_service_id"] = user["id"]
    
    # Filter by status
    if status:
        query["status"] = status
    
    # NEW: Filter by towing service (for authorities)
    if service_id and user["role"] == UserRole.AUTHORITY:
        query["assigned_service_id"] = service_id
    
    # Filter by date range
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            # Add time to include the entire day
            date_query["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_query
    
    # Search
    if search:
        search_upper = search.upper()
        query["$or"] = [
            {"license_plate": {"$regex": search_upper, "$options": "i"}},
            {"vin": {"$regex": search_upper, "$options": "i"}},
            {"job_number": {"$regex": search_upper, "$options": "i"}}
        ]
    
    # Pagination
    skip = (page - 1) * limit
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return [JobResponse(**j) for j in jobs]

@api_router.get("/jobs/count/total")
async def get_jobs_count(
    status: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    service_id: Optional[str] = None,  # NEW: Filter by towing service
    user: dict = Depends(get_current_user)
):
    """Get total count of jobs for pagination"""
    query = {}
    
    # Filter by role
    if user["role"] == UserRole.AUTHORITY:
        if user.get("is_main_authority"):
            query["authority_id"] = user["id"]
        else:
            query["created_by_id"] = user["id"]
    elif user["role"] == UserRole.TOWING_SERVICE:
        query["assigned_service_id"] = user["id"]
    
    if status:
        query["status"] = status
    
    # NEW: Filter by towing service (for authorities)
    if service_id and user["role"] == UserRole.AUTHORITY:
        query["assigned_service_id"] = service_id
    
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_query
    
    if search:
        search_upper = search.upper()
        query["$or"] = [
            {"license_plate": {"$regex": search_upper, "$options": "i"}},
            {"vin": {"$regex": search_upper, "$options": "i"}},
            {"job_number": {"$regex": search_upper, "$options": "i"}}
        ]
    
    count = await db.jobs.count_documents(query)
    return {"total": count}

@api_router.get("/jobs/updates")
async def get_jobs_updates(
    since: str,
    user: dict = Depends(get_current_user)
):
    """Delta polling endpoint: Returns jobs updated since the given ISO timestamp"""
    query = {"updated_at": {"$gt": since}}
    
    # Filter by role
    if user["role"] == UserRole.AUTHORITY:
        if user.get("is_main_authority"):
            query["authority_id"] = user["id"]
        else:
            query["created_by_id"] = user["id"]
    elif user["role"] == UserRole.TOWING_SERVICE:
        query["assigned_service_id"] = user["id"]
        
    jobs = await db.jobs.find(query, {"_id": 0}).sort("updated_at", 1).limit(500).to_list(1000)
    current_time = datetime.now(timezone.utc).isoformat()
    
    return {
        "serverNow": current_time,
        "changedOrders": [JobResponse(**j) for j in jobs]
    }

@api_router.get("/jobs/{job_id}", response_model=JobResponse)
async def get_job(job_id: str, user: dict = Depends(get_current_user)):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Check access for authority users
    if user["role"] == UserRole.AUTHORITY:
        authority_id = get_authority_id(user)
        if job.get("authority_id") != authority_id and job["created_by_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
    if user["role"] == UserRole.TOWING_SERVICE and job["assigned_service_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return JobResponse(**job)

@api_router.patch("/jobs/{job_id}", response_model=JobResponse)
async def update_job(job_id: str, data: JobUpdate, user: dict = Depends(get_current_user)):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Check access for authority users
    if user["role"] == UserRole.AUTHORITY:
        authority_id = get_authority_id(user)
        if job.get("authority_id") != authority_id and job["created_by_id"] != user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
    if user["role"] == UserRole.TOWING_SERVICE and job["assigned_service_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = datetime.now(timezone.utc).isoformat()
    update_data = {"updated_at": now}
    
    if data.status:
        update_data["status"] = data.status
        if data.status == JobStatus.ON_SITE:
            update_data["on_site_at"] = now
            # Clear future steps
            update_data["towed_at"] = None
            update_data["in_yard_at"] = None
            update_data["delivered_to_authority_at"] = None
            update_data["released_at"] = None
            # Set accepted_at when service first accepts the job
            if not job.get("accepted_at"):
                update_data["accepted_at"] = now
        elif data.status == JobStatus.TOWED:
            update_data["towed_at"] = now
            # Clear future steps
            update_data["in_yard_at"] = None
            update_data["delivered_to_authority_at"] = None
            update_data["released_at"] = None
        elif data.status == JobStatus.IN_YARD:
            # Check if this is authority_yard - if so, use DELIVERED_TO_AUTHORITY status instead
            if job.get("target_yard") == "authority_yard":
                update_data["status"] = JobStatus.DELIVERED_TO_AUTHORITY
                update_data["delivered_to_authority_at"] = now
            else:
                update_data["in_yard_at"] = now
            # Clear future steps
            update_data["released_at"] = None
        elif data.status == JobStatus.DELIVERED_TO_AUTHORITY:
            update_data["delivered_to_authority_at"] = now
            # Clear future steps
            update_data["released_at"] = None
        elif data.status == JobStatus.RELEASED:
            update_data["released_at"] = now
        elif data.status == JobStatus.ASSIGNED:
            # Clear all forward steps
            update_data["on_site_at"] = None
            update_data["towed_at"] = None
            update_data["in_yard_at"] = None
            update_data["delivered_to_authority_at"] = None
            update_data["released_at"] = None
    
    if data.is_empty_trip is not None:
        update_data["is_empty_trip"] = data.is_empty_trip
    
    if data.photos:
        update_data["photos"] = data.photos
    if data.notes is not None:
        update_data["notes"] = data.notes
    if data.service_notes is not None:
        update_data["service_notes"] = data.service_notes
    if data.owner_first_name is not None:
        update_data["owner_first_name"] = data.owner_first_name
    if data.owner_last_name is not None:
        update_data["owner_last_name"] = data.owner_last_name
    if data.owner_address is not None:
        update_data["owner_address"] = data.owner_address
    if data.payment_method is not None:
        update_data["payment_method"] = data.payment_method
    if data.payment_amount is not None:
        update_data["payment_amount"] = data.payment_amount
    
    await db.jobs.update_one({"id": job_id}, {"$set": update_data})
    
    # Audit log job update
    if data.status:
        await log_audit("JOB_STATUS_UPDATED", user["id"], user.get("name", user["email"]), {
            "job_id": job_id,
            "job_number": job.get("job_number"),
            "license_plate": job.get("license_plate"),
            "old_status": job.get("status"),
            "new_status": data.status
        })
    
    updated_job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    return JobResponse(**updated_job)

# ==================== AUTHORITY YARD RELEASE ====================

class AuthorityReleaseData(BaseModel):
    owner_first_name: str
    owner_last_name: str
    owner_address: Optional[str] = None
    payment_method: str
    payment_amount: float
    service_invoice_amount: Optional[float] = None  # Amount to pay towing service

@api_router.post("/jobs/{job_id}/authority-release", response_model=JobResponse)
async def authority_release_vehicle(job_id: str, data: AuthorityReleaseData, user: dict = Depends(get_current_user)):
    """Release vehicle from authority yard - only authorities can do this"""
    if user["role"] != UserRole.AUTHORITY:
        raise HTTPException(status_code=403, detail="Nur Behörden können Fahrzeuge aus dem Behörden-Hof freigeben")
    
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Auftrag nicht gefunden")
    
    # Verify this is an authority yard job
    if job.get("target_yard") != "authority_yard":
        raise HTTPException(status_code=400, detail="Dieser Auftrag ist nicht für den Behörden-Hof bestimmt")
    
    # Verify job is in correct status
    if job.get("status") != JobStatus.DELIVERED_TO_AUTHORITY:
        raise HTTPException(status_code=400, detail="Das Fahrzeug wurde noch nicht an die Behörde übergeben")
    
    # Check authority has access
    authority_id = get_authority_id(user)
    if job.get("authority_id") != authority_id:
        raise HTTPException(status_code=403, detail="Kein Zugriff auf diesen Auftrag")
    
    now = datetime.now(timezone.utc).isoformat()
    
    update_data = {
        "status": JobStatus.RELEASED,
        "released_at": now,
        "updated_at": now,
        "owner_first_name": data.owner_first_name,
        "owner_last_name": data.owner_last_name,
        "owner_address": data.owner_address,
        "payment_method": data.payment_method,
        "payment_amount": data.payment_amount,
        # Store service invoice info
        "service_invoice_amount": data.service_invoice_amount,
        "service_invoice_created_at": now if data.service_invoice_amount else None
    }
    
    await db.jobs.update_one({"id": job_id}, {"$set": update_data})
    
    # Create invoice record for towing service (if service_invoice_amount is set)
    if data.service_invoice_amount and job.get("assigned_service_id"):
        invoice_doc = {
            "id": str(uuid.uuid4()),
            "job_id": job_id,
            "job_number": job.get("job_number"),
            "license_plate": job.get("license_plate"),
            "authority_id": authority_id,
            "authority_name": user.get("authority_name", user.get("name")),
            "service_id": job.get("assigned_service_id"),
            "service_name": job.get("assigned_service_name"),
            "amount": data.service_invoice_amount,
            "status": "pending",  # pending, paid
            "created_at": now,
            "paid_at": None
        }
        await db.service_invoices.insert_one(invoice_doc)
    
    # Audit log
    await log_audit("AUTHORITY_RELEASE", user["id"], user["name"], {
        "job_id": job_id,
        "job_number": job.get("job_number"),
        "payment_amount": data.payment_amount,
        "service_invoice_amount": data.service_invoice_amount
    })
    
    updated_job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    return JobResponse(**updated_job)

# NEW: Get invoices for towing service
@api_router.get("/services/invoices")
async def get_service_invoices(user: dict = Depends(get_current_user)):
    """Get pending and paid invoices for a towing service"""
    if user["role"] != UserRole.TOWING_SERVICE:
        raise HTTPException(status_code=403, detail="Nur für Abschleppdienste")
    
    invoices = await db.service_invoices.find(
        {"service_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return {"invoices": invoices}

# ==================== JOB DATA EDITING (Kennzeichen, FIN, etc.) ====================

@api_router.patch("/jobs/{job_id}/edit-data", response_model=JobResponse)
async def edit_job_data(job_id: str, data: JobEditData, user: dict = Depends(get_current_user)):
    """Edit job vehicle data - allows corrections to license plate, VIN, etc."""
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Auftrag nicht gefunden")
    
    # Check access permissions
    has_access = False
    
    if user["role"] == UserRole.ADMIN:
        has_access = True
    elif user["role"] == UserRole.AUTHORITY:
        authority_id = get_authority_id(user)
        if job.get("authority_id") == authority_id or job["created_by_id"] == user["id"]:
            has_access = True
    elif user["role"] == UserRole.TOWING_SERVICE:
        if job["assigned_service_id"] == user["id"]:
            has_access = True
    
    if not has_access:
        raise HTTPException(status_code=403, detail="Keine Berechtigung zum Bearbeiten dieses Auftrags")
    
    # Only allow editing if job is not released
    if job["status"] == JobStatus.RELEASED:
        raise HTTPException(status_code=400, detail="Abgeschlossene Aufträge können nicht mehr bearbeitet werden")
    
    now = datetime.now(timezone.utc).isoformat()
    update_data = {"updated_at": now}
    changes = []
    
    # Update license plate if provided and different
    if data.license_plate and data.license_plate.upper() != job.get("license_plate"):
        new_plate = data.license_plate.upper()
        
        # Check if new license plate already exists in another active job
        normalized_new_plate = new_plate.replace(" ", "").replace("-", "")
        existing_job = await db.jobs.find_one({
            "id": {"$ne": job_id},  # Exclude current job
            "$or": [
                {"license_plate": new_plate},
                {"license_plate": normalized_new_plate}
            ],
            "status": {"$ne": JobStatus.RELEASED}
        })
        
        if existing_job:
            raise HTTPException(
                status_code=400, 
                detail=f"Ein Fahrzeug mit diesem Kennzeichen ({new_plate}) existiert bereits im System"
            )
        
        changes.append(f"Kennzeichen: {job.get('license_plate')} → {new_plate}")
        update_data["license_plate"] = new_plate
    
    # Update VIN if provided
    if data.vin is not None:
        new_vin = data.vin.upper() if data.vin else None
        if new_vin != job.get("vin"):
            changes.append(f"FIN: {job.get('vin') or '(leer)'} → {new_vin or '(leer)'}")
            update_data["vin"] = new_vin
    
    # Update tow reason if provided
    if data.tow_reason and data.tow_reason != job.get("tow_reason"):
        changes.append(f"Abschleppgrund: {job.get('tow_reason')} → {data.tow_reason}")
        update_data["tow_reason"] = data.tow_reason
    
    # Update notes if provided
    if data.notes is not None and data.notes != job.get("notes"):
        changes.append("Bemerkungen aktualisiert")
        update_data["notes"] = data.notes
    
    # Update location if provided
    if data.location_address and data.location_address != job.get("location_address"):
        changes.append(f"Adresse aktualisiert")
        update_data["location_address"] = data.location_address
    if data.location_lat is not None:
        update_data["location_lat"] = data.location_lat
    if data.location_lng is not None:
        update_data["location_lng"] = data.location_lng
    
    if len(update_data) <= 1:  # Only updated_at
        raise HTTPException(status_code=400, detail="Keine Änderungen vorgenommen")
    
    await db.jobs.update_one({"id": job_id}, {"$set": update_data})
    
    # Audit log the changes
    await log_audit("JOB_DATA_EDITED", user["id"], user.get("name", user["email"]), {
        "job_id": job_id,
        "job_number": job.get("job_number"),
        "changes": changes
    })
    
    updated_job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    return JobResponse(**updated_job)

# ==================== DELETE JOB ====================

@api_router.delete("/jobs/{job_id}")
async def delete_job(job_id: str, user: dict = Depends(get_current_user)):
    """Delete a job - only allowed for creator before it's too far in process"""
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Auftrag nicht gefunden")
    
    # Check access permissions - only creator or admin can delete
    has_access = False
    
    if user["role"] == UserRole.ADMIN:
        has_access = True
    elif user["role"] == UserRole.AUTHORITY:
        authority_id = get_authority_id(user)
        # Authority can delete if they created the job
        if job.get("authority_id") == authority_id or job["created_by_id"] == user["id"]:
            has_access = True
    elif user["role"] == UserRole.TOWING_SERVICE:
        # Towing service can delete if they created the job (for_authority jobs)
        if job.get("created_by_id") == user["id"]:
            has_access = True
    
    if not has_access:
        raise HTTPException(status_code=403, detail="Keine Berechtigung zum Löschen dieses Auftrags")
    
    # Only allow deletion if job is in early stages (pending, assigned, or on_site)
    # Don't allow deletion if vehicle is already towed/in_yard/released
    allowed_delete_statuses = [JobStatus.PENDING, JobStatus.ASSIGNED, JobStatus.ON_SITE]
    if job["status"] not in allowed_delete_statuses:
        raise HTTPException(
            status_code=400, 
            detail=f"Auftrag kann nicht gelöscht werden - Fahrzeug wurde bereits abgeschleppt. Status: {job['status']}"
        )
    
    # Delete the job
    await db.jobs.delete_one({"id": job_id})
    
    # Audit log the deletion
    await log_audit("JOB_DELETED", user["id"], user.get("name", user["email"]), {
        "job_id": job_id,
        "job_number": job.get("job_number"),
        "license_plate": job.get("license_plate"),
        "status_at_deletion": job["status"],
        "reason": "User requested deletion"
    })
    
    return {"message": f"Auftrag {job['job_number']} wurde gelöscht"}

# ==================== BULK STATUS UPDATE ====================

@api_router.post("/jobs/bulk-update-status")
async def bulk_update_status(data: BulkStatusUpdate, user: dict = Depends(get_current_user)):
    """Bulk update status for multiple jobs at once"""
    if user["role"] != UserRole.TOWING_SERVICE:
        raise HTTPException(status_code=403, detail="Only towing services can bulk update jobs")
    
    if not data.job_ids:
        raise HTTPException(status_code=400, detail="Keine Aufträge ausgewählt")
    
    valid_statuses = [JobStatus.ON_SITE, JobStatus.TOWED, JobStatus.IN_YARD]
    if data.status not in valid_statuses:
        raise HTTPException(status_code=400, detail="Ungültiger Status für Massenänderung")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Build update data based on status
    update_data = {"status": data.status, "updated_at": now}
    if data.status == JobStatus.ON_SITE:
        update_data["on_site_at"] = now
    elif data.status == JobStatus.TOWED:
        update_data["towed_at"] = now
    elif data.status == JobStatus.IN_YARD:
        update_data["in_yard_at"] = now
    
    # Only update jobs that belong to this towing service
    result = await db.jobs.update_many(
        {"id": {"$in": data.job_ids}, "assigned_service_id": user["id"]},
        {"$set": update_data}
    )
    
    return {
        "message": f"{result.modified_count} Aufträge aktualisiert",
        "updated_count": result.modified_count
    }

@api_router.post("/jobs/{job_id}/assign/{service_id}", response_model=JobResponse)
async def assign_job(job_id: str, service_id: str, user: dict = Depends(get_current_user)):
    if user["role"] not in [UserRole.AUTHORITY, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Only authorities can assign jobs")
    
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    service = await db.users.find_one({"id": service_id, "role": UserRole.TOWING_SERVICE})
    if not service:
        raise HTTPException(status_code=404, detail="Towing service not found")
    
    now = datetime.now(timezone.utc).isoformat()
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {
            "assigned_service_id": service_id,
            "assigned_service_name": service["company_name"],
            "status": JobStatus.ASSIGNED,
            "updated_at": now
        }}
    )
    
    updated_job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    return JobResponse(**updated_job)

@api_router.post("/jobs/{job_id}/photos")
async def add_service_photo(job_id: str, photo: str = Form(...), user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.TOWING_SERVICE:
        raise HTTPException(status_code=403, detail="Only towing services can add photos")
    
    job = await db.jobs.find_one({"id": job_id})
    if not job or job["assigned_service_id"] != user["id"]:
        raise HTTPException(status_code=404, detail="Job not found")
    
    await db.jobs.update_one(
        {"id": job_id},
        {"$push": {"service_photos": photo}}
    )
    
    return {"message": "Photo added"}

# ==================== PUBLIC SEARCH ====================

@api_router.get("/search/vehicle", response_model=VehicleSearchResult)
async def search_vehicle(q: str):
    search_term = q.upper().strip()
    
    job = await db.jobs.find_one({
        "$or": [
            {"license_plate": search_term},
            {"vin": search_term}
        ],
        "status": {"$in": [JobStatus.TOWED, JobStatus.IN_YARD, JobStatus.DELIVERED_TO_AUTHORITY]}
    }, {"_id": 0})
    
    if not job:
        return VehicleSearchResult(found=False)
    
    # Check if this is an authority yard job
    is_authority_yard = job.get("target_yard") == "authority_yard"
    
    # Initialize variables
    service = None
    authority = None
    tow_cost = None
    daily_cost = None
    processing_fee = None
    night_surcharge = None
    days_in_yard = 0
    total_cost = None
    
    yard_address = None
    yard_lat = None
    yard_lng = None
    company_name = None
    phone = None
    email = None
    opening_hours = None
    
    if is_authority_yard:
        # AUTHORITY YARD: Get authority info and pricing
        if job.get("authority_id"):
            authority = await db.users.find_one({"id": job["authority_id"]}, {"_id": 0, "password": 0})
        
        if authority:
            # Use authority yard location from job
            yard_address = job.get("authority_yard_address") or authority.get("yard_address")
            yard_lat = job.get("authority_yard_lat") or authority.get("yard_lat")
            yard_lng = job.get("authority_yard_lng") or authority.get("yard_lng")
            phone = job.get("authority_yard_phone") or authority.get("phone")
            company_name = job.get("authority_yard_name") or authority.get("authority_name") or "Behörden-Hof"
            email = authority.get("email")
            
            # Use authority pricing
            tow_cost = job.get("authority_base_price") or 0
            daily_cost = job.get("authority_daily_rate") or 0
            processing_fee = 0  # Authority may not have processing fee
            
            # Calculate days in yard
            if job.get("delivered_to_authority_at"):
                days_in_yard = calculate_days_in_yard(job["delivered_to_authority_at"])
            elif job.get("in_yard_at"):
                days_in_yard = calculate_days_in_yard(job["in_yard_at"])
            elif job.get("towed_at"):
                days_in_yard = calculate_days_in_yard(job["towed_at"])
            
            # Calculate total cost (authority pricing)
            standkosten = daily_cost * max(0, days_in_yard - 1)  # First day included in base
            total_cost = tow_cost + standkosten
    else:
        # SERVICE YARD: Get towing service info (original logic)
        if job.get("assigned_service_id"):
            service = await db.users.find_one({"id": job["assigned_service_id"]}, {"_id": 0, "password": 0})
            if service:
                tow_cost = service.get("tow_cost", 0) or 0
                daily_cost = service.get("daily_cost", 0) or 0
                processing_fee = service.get("processing_fee", 0) or 0
                empty_trip_fee = 0
                heavy_vehicle_surcharge = 0
                night_surcharge = 0
                weekend_surcharge = 0
                
                yard_address = service.get("yard_address")
                yard_lat = service.get("yard_lat")
                yard_lng = service.get("yard_lng")
                company_name = service.get("company_name")
                phone = service.get("phone")
                email = service.get("email")
                opening_hours = service.get("opening_hours")
                
                # Empty trip fee
                if job.get("is_empty_trip"):
                    empty_trip_fee = service.get("empty_trip_fee", 0) or 0
                    
                # Heavy vehicle surcharge (or weight category surcharge)
                if job.get("weight_category_surcharge"):
                    heavy_vehicle_surcharge = job.get("weight_category_surcharge", 0) or 0
                elif job.get("vehicle_category") == "over_3_5t":
                    heavy_vehicle_surcharge = service.get("heavy_vehicle_surcharge", 0) or 0
                    
                # Night and Weekend surcharges
                if job.get("towed_at"):
                    towed_time = datetime.fromisoformat(job["towed_at"].replace("Z", "+00:00"))
                    hour = towed_time.hour
                    if hour >= 22 or hour < 6:
                        night_surcharge = service.get("night_surcharge", 0) or 0
                    if towed_time.weekday() >= 5:
                        weekend_surcharge = service.get("weekend_surcharge", 0) or 0
                
                # Calculate days in yard
                if job.get("in_yard_at"):
                    days_in_yard = calculate_days_in_yard(job["in_yard_at"])
                elif job.get("towed_at"):
                    days_in_yard = calculate_days_in_yard(job["towed_at"])
                
                # Calculate total cost including all fees
                standkosten = daily_cost * days_in_yard
                total_cost = tow_cost + standkosten + processing_fee + empty_trip_fee + heavy_vehicle_surcharge + night_surcharge + weekend_surcharge
    
    # Map status for display
    display_status = job["status"]
    if display_status == "delivered_to_authority":
        display_status = "in_yard"  # Show as "Im Hof" for public search
    
    return VehicleSearchResult(
        found=True,
        job_number=job["job_number"],
        license_plate=job["license_plate"],
        status=display_status,
        towed_at=job.get("towed_at"),
        in_yard_at=job.get("delivered_to_authority_at") or job.get("in_yard_at"),
        yard_address=yard_address,
        yard_lat=yard_lat,
        yard_lng=yard_lng,
        company_name=company_name,
        phone=phone,
        email=email,
        opening_hours=opening_hours,
        tow_cost=tow_cost,
        daily_cost=daily_cost,
        days_in_yard=days_in_yard,
        processing_fee=processing_fee,
        empty_trip_fee=empty_trip_fee if 'empty_trip_fee' in locals() else None,
        night_surcharge=night_surcharge,
        weekend_surcharge=weekend_surcharge if 'weekend_surcharge' in locals() else None,
        heavy_vehicle_surcharge=heavy_vehicle_surcharge if 'heavy_vehicle_surcharge' in locals() else None,
        total_cost=total_cost,
        location_lat=job.get("location_lat"),
        location_lng=job.get("location_lng"),
        tow_reason=job.get("tow_reason"),
        location_address=job.get("location_address"),
        created_by_authority=job.get("created_by_authority")
    )

# ==================== ADMIN ROUTES ====================

@api_router.get("/admin/stats", response_model=StatsResponse)
async def get_admin_stats(user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    total_jobs = await db.jobs.count_documents({})
    pending_jobs = await db.jobs.count_documents({"status": {"$in": [JobStatus.PENDING, JobStatus.ASSIGNED, JobStatus.ON_SITE, JobStatus.TOWED]}})
    in_yard_jobs = await db.jobs.count_documents({"status": JobStatus.IN_YARD})
    released_jobs = await db.jobs.count_documents({"status": JobStatus.RELEASED})
    total_services = await db.users.count_documents({"role": UserRole.TOWING_SERVICE, "approval_status": ApprovalStatus.APPROVED})
    total_authorities = await db.users.count_documents({"role": UserRole.AUTHORITY})
    pending_approvals = await db.users.count_documents({"role": UserRole.TOWING_SERVICE, "approval_status": ApprovalStatus.PENDING})
    
    return StatsResponse(
        total_jobs=total_jobs,
        pending_jobs=pending_jobs,
        in_yard_jobs=in_yard_jobs,
        released_jobs=released_jobs,
        total_services=total_services,
        total_authorities=total_authorities,
        pending_approvals=pending_approvals
    )

@api_router.get("/admin/jobs", response_model=List[JobResponse])
async def get_all_jobs(
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    query = {}
    if status:
        query["status"] = status
    if search:
        # Full-text search across multiple fields
        query["$or"] = [
            {"license_plate": {"$regex": search, "$options": "i"}},
            {"vin": {"$regex": search, "$options": "i"}},
            {"job_number": {"$regex": search, "$options": "i"}},
            {"tow_reason": {"$regex": search, "$options": "i"}},
            {"notes": {"$regex": search, "$options": "i"}},
            {"service_notes": {"$regex": search, "$options": "i"}},
            {"location_address": {"$regex": search, "$options": "i"}},
            {"created_by_name": {"$regex": search, "$options": "i"}},
            {"assigned_service_name": {"$regex": search, "$options": "i"}},
            {"owner_first_name": {"$regex": search, "$options": "i"}},
            {"owner_last_name": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * limit
    jobs = await db.jobs.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    result = []
    for j in jobs:
        # Ensure id field exists
        if "id" not in j and "_id" in j:
            j["id"] = str(j.pop("_id"))
        elif "_id" in j:
            j.pop("_id")
        result.append(JobResponse(**j))
    return result

@api_router.get("/admin/jobs/count")
async def get_all_jobs_count(
    status: Optional[str] = None,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get total count of jobs for admin pagination"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    query = {}
    if status:
        query["status"] = status
    if search:
        # Full-text search across multiple fields
        query["$or"] = [
            {"license_plate": {"$regex": search, "$options": "i"}},
            {"vin": {"$regex": search, "$options": "i"}},
            {"job_number": {"$regex": search, "$options": "i"}},
            {"tow_reason": {"$regex": search, "$options": "i"}},
            {"notes": {"$regex": search, "$options": "i"}},
            {"service_notes": {"$regex": search, "$options": "i"}},
            {"location_address": {"$regex": search, "$options": "i"}},
            {"created_by_name": {"$regex": search, "$options": "i"}},
            {"assigned_service_name": {"$regex": search, "$options": "i"}},
            {"owner_first_name": {"$regex": search, "$options": "i"}},
            {"owner_last_name": {"$regex": search, "$options": "i"}}
        ]
    
    count = await db.jobs.count_documents(query)
    return {"total": count}

@api_router.get("/admin/users", response_model=List[UserResponse])
async def get_all_users(user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    users = await db.users.find({}, {"password": 0}).to_list(1000)
    result = []
    for u in users:
        # Ensure id field exists
        if "id" not in u and "_id" in u:
            u["id"] = str(u.pop("_id"))
        elif "_id" in u:
            u.pop("_id")
        result.append(UserResponse(**u))
    return result

# ==================== ADMIN USER MANAGEMENT ====================

@api_router.patch("/admin/users/{user_id}/password")
async def admin_update_password(user_id: str, data: AdminUpdatePasswordRequest, user: dict = Depends(get_current_user)):
    """Admin can update any user's password"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Benutzer nicht gefunden")
    
    # Don't allow changing own password through this endpoint
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Eigenes Passwort bitte über Profil ändern")
    
    new_hashed = hash_password(data.new_password)
    await db.users.update_one({"id": user_id}, {"$set": {"password": new_hashed}})
    
    # Audit log password change by admin
    await log_audit("ADMIN_PASSWORD_CHANGE", user["id"], user["name"], {
        "target_user_id": user_id,
        "target_user_email": target_user.get("email"),
        "target_user_name": target_user.get("name")
    })
    
    return {"message": f"Passwort für {target_user.get('name', target_user['email'])} wurde aktualisiert"}

@api_router.patch("/admin/users/{user_id}/block")
async def admin_block_user(user_id: str, data: AdminBlockUserRequest, user: dict = Depends(get_current_user)):
    """Admin can block/unblock users"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Benutzer nicht gefunden")
    
    # Don't allow blocking yourself
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Sie können sich nicht selbst sperren")
    
    # Don't allow blocking other admins
    if target_user["role"] == UserRole.ADMIN:
        raise HTTPException(status_code=400, detail="Administratoren können nicht gesperrt werden")
    
    await db.users.update_one({"id": user_id}, {"$set": {"is_blocked": data.blocked}})
    
    action = "gesperrt" if data.blocked else "entsperrt"
    
    # Audit log user block/unblock
    await log_audit("USER_BLOCKED" if data.blocked else "USER_UNBLOCKED", user["id"], user["name"], {
        "target_user_id": user_id,
        "target_user_email": target_user.get("email"),
        "target_user_name": target_user.get("name"),
        "blocked": data.blocked
    })
    
    return {"message": f"{target_user.get('name', target_user['email'])} wurde {action}"}

@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, user: dict = Depends(get_current_user)):
    """Admin can permanently delete users"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Benutzer nicht gefunden")
    
    # Don't allow deleting yourself
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Sie können sich nicht selbst löschen")
    
    # Don't allow deleting other admins
    if target_user["role"] == UserRole.ADMIN:
        raise HTTPException(status_code=400, detail="Administratoren können nicht gelöscht werden")
    
    user_name = target_user.get('name', target_user['email'])
    
    # Delete the user
    await db.users.delete_one({"id": user_id})
    
    # Audit log user deletion
    await log_audit("USER_DELETED", user["id"], user["name"], {
        "deleted_user_id": user_id,
        "deleted_user_email": target_user.get("email"),
        "deleted_user_name": user_name,
        "deleted_user_role": target_user.get("role")
    })
    
    return {"message": f"{user_name} wurde permanent gelöscht"}

# ==================== PDF GENERATION ====================

def format_datetime(dt_str):
    """Format ISO datetime string to German format"""
    if not dt_str or dt_str == '-':
        return '-'
    try:
        dt = datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
        return dt.strftime('%d.%m.%Y %H:%M')
    except:
        return dt_str[:19].replace('T', ' ')

def wrap_text(text, max_length=60):
    """Wrap long text for PDF cells"""
    if not text:
        return '-'
    text = str(text)
    if len(text) <= max_length:
        return text
    # Insert line breaks for very long text
    words = text.split()
    lines = []
    current_line = ""
    for word in words:
        if len(current_line + " " + word) <= max_length:
            current_line = (current_line + " " + word).strip()
        else:
            if current_line:
                lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)
    return "\n".join(lines)

@api_router.get("/jobs/{job_id}/pdf/token")
async def get_pdf_download_token(job_id: str, user: dict = Depends(get_current_user)):
    """Generate a short-lived token to download a PDF (requires auth)"""
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
        
    # Check permissions
    if user["role"] == UserRole.AUTHORITY:
        if user.get("is_main_authority"):
            if job.get("authority_id") != user["id"]:
                raise HTTPException(status_code=403, detail="Keine Berechtigung für diesen Auftrag")
        else:
            if job.get("created_by_id") != user["id"] and job.get("authority_id") != user.get("parent_authority_id"):
                raise HTTPException(status_code=403, detail="Keine Berechtigung für diesen Auftrag")
    elif user["role"] == UserRole.TOWING_SERVICE:
        if job.get("assigned_service_id") != user["id"]:
            raise HTTPException(status_code=403, detail="Keine Berechtigung für diesen Auftrag")
            
    # Generate 5-minute token
    expiration = datetime.now(timezone.utc) + timedelta(minutes=5)
    payload = {
        "job_id": job_id,
        "action": "download_pdf",
        "exp": expiration
    }
    
    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return {"token": token}

@api_router.get("/jobs/{job_id}/pdf")
async def generate_pdf(job_id: str, token: str):
    """Generate PDF - secured by short-lived token"""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("job_id") != job_id or payload.get("action") != "download_pdf":
            raise HTTPException(status_code=403, detail="Ungültiger Token")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=403, detail="Token abgelaufen (nur 5 min gültig)")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=403, detail="Ungültiger Token")
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get service info
    service = None
    if job.get("assigned_service_id"):
        service = await db.users.find_one({"id": job["assigned_service_id"]}, {"_id": 0, "password": 0})
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        rightMargin=1.5*cm, 
        leftMargin=1.5*cm, 
        topMargin=1.5*cm, 
        bottomMargin=1.5*cm
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle', 
        parent=styles['Heading1'], 
        fontSize=20, 
        spaceAfter=5,
        alignment=1,  # Center
        textColor=colors.HexColor('#1e293b')
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', 
        parent=styles['Normal'], 
        fontSize=11, 
        spaceAfter=20,
        alignment=1,  # Center
        textColor=colors.HexColor('#64748b')
    )
    heading_style = ParagraphStyle(
        'CustomHeading', 
        parent=styles['Heading2'], 
        fontSize=12, 
        spaceAfter=8, 
        spaceBefore=15,
        textColor=colors.HexColor('#1e293b'),
        borderPadding=(0, 0, 5, 0)
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        wordWrap='CJK'
    )
    
    story = []
    
    # ===== HEADER =====
    # Check if this is an empty trip
    is_empty_trip = job.get('is_empty_trip', False)
    
    if is_empty_trip:
        story.append(Paragraph("LEERFAHRT-PROTOKOLL", title_style))
    else:
        story.append(Paragraph("ABSCHLEPPPROTOKOLL", title_style))
    story.append(Paragraph(f"Auftragsnummer: {job['job_number']}", subtitle_style))
    
    # Horizontal line
    story.append(Spacer(1, 5))
    line_table = Table([['']], colWidths=[17*cm])
    line_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.HexColor('#f97316')),
    ]))
    story.append(line_table)
    story.append(Spacer(1, 10))
    
    # ===== EMPTY TRIP DETAILS (if applicable) =====
    if is_empty_trip:
        story.append(Paragraph("Leerfahrt-Details", heading_style))
        
        # Parse service notes for empty trip reason
        service_notes = job.get('service_notes', '')
        empty_trip_reason = 'Fahrzeug war nicht mehr vor Ort'
        if 'Fahrer vor Ort angetroffen' in service_notes:
            empty_trip_reason = 'Fahrer vor Ort angetroffen'
        
        empty_data = [
            [Paragraph("<b>Status</b>", cell_style), Paragraph("<b>LEERFAHRT</b>", cell_style)],
            [Paragraph("<b>Grund</b>", cell_style), Paragraph(empty_trip_reason, cell_style)],
        ]
        
        # Add driver info if present
        if job.get('owner_first_name') and job.get('owner_last_name'):
            empty_data.append([
                Paragraph("<b>Angetroffene Person</b>", cell_style), 
                Paragraph(f"{job['owner_first_name']} {job['owner_last_name']}", cell_style)
            ])
            if job.get('owner_address'):
                empty_data.append([
                    Paragraph("<b>Adresse</b>", cell_style), 
                    Paragraph(job['owner_address'], cell_style)
                ])
        
        # Payment info
        if job.get('payment_amount'):
            payment_method_labels = {'cash': 'Bar', 'card': 'Karte', 'invoice': 'Rechnung'}
            payment_method = payment_method_labels.get(job.get('payment_method', ''), job.get('payment_method', '-'))
            empty_data.append([
                Paragraph("<b>Leerfahrt-Kosten</b>", cell_style), 
                Paragraph(f"{job['payment_amount']:.2f} €", cell_style)
            ])
            empty_data.append([
                Paragraph("<b>Zahlungsart</b>", cell_style), 
                Paragraph(payment_method, cell_style)
            ])
        
        empty_table = Table(empty_data, colWidths=[5*cm, 12*cm])
        empty_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fff7ed')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#fed7aa')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fdba74')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(empty_table)
        story.append(Spacer(1, 15))
    
    # ===== AUFTRAGSTYP =====
    job_type = job.get('job_type', 'towing')
    job_type_label = 'Sicherstellung (Polizeilich)' if job_type == 'sicherstellung' else 'Abschleppen (Falschparker)'
    
    story.append(Paragraph("Auftragsdetails", heading_style))
    order_data = [
        [Paragraph("<b>Auftragstyp</b>", cell_style), Paragraph(f"<b>{job_type_label}</b>", cell_style)],
    ]
    
    # Sicherstellung-specific details
    if job_type == 'sicherstellung':
        # Reason mapping
        reason_labels = {
            'betriebsmittel': 'Auslaufende Betriebsmittel',
            'gestohlen': 'Gestohlenes Fahrzeug / Fahndung',
            'eigentumssicherung': 'Eigentumssicherung (wertvoll/ungesichert)',
            'technische_maengel': 'Technische Mängel / Beweissicherung',
            'strafrechtlich': 'Strafrechtliche Beschlagnahme'
        }
        sicherstellung_reason = job.get('sicherstellung_reason', '')
        reason_text = reason_labels.get(sicherstellung_reason, sicherstellung_reason or '-')
        order_data.append([
            Paragraph("<b>Grund der Sicherstellung</b>", cell_style), 
            Paragraph(reason_text, cell_style)
        ])
        
        # Vehicle category
        vehicle_cat = job.get('vehicle_category', '')
        cat_label = 'PKW/Krad bis 3,5t' if vehicle_cat == 'under_3_5t' else 'Fahrzeuge ab 3,5t' if vehicle_cat == 'over_3_5t' else '-'
        order_data.append([
            Paragraph("<b>Fahrzeugkategorie</b>", cell_style), 
            Paragraph(cat_label, cell_style)
        ])
        
        # Ordering authority
        authority_labels = {
            'schutzpolizei': 'Schutzpolizei',
            'kriminalpolizei': 'Kriminalpolizei',
            'staatsanwaltschaft': 'Staatsanwaltschaft',
            'sachverstaendiger': 'Technischer Sachverständiger'
        }
        ordering_auth = job.get('ordering_authority', '')
        auth_text = authority_labels.get(ordering_auth, ordering_auth or '-')
        order_data.append([
            Paragraph("<b>Anordnende Stelle</b>", cell_style), 
            Paragraph(auth_text, cell_style)
        ])
        
        # Contact attempts
        contact_attempts = job.get('contact_attempts')
        if contact_attempts is not None:
            contact_text = 'Ja' if contact_attempts else 'Nein'
            if contact_attempts and job.get('contact_attempts_notes'):
                contact_text += f" - {job['contact_attempts_notes']}"
            order_data.append([
                Paragraph("<b>Telefonische Kontaktversuche</b>", cell_style), 
                Paragraph(wrap_text(contact_text, 50), cell_style)
            ])
        
        # Estimated vehicle value
        if job.get('estimated_vehicle_value'):
            order_data.append([
                Paragraph("<b>Geschätzter Fahrzeugwert</b>", cell_style), 
                Paragraph(f"{job['estimated_vehicle_value']:,.2f} €", cell_style)
            ])
    
    order_table = Table(order_data, colWidths=[4.5*cm, 12.5*cm])
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fef3c7') if job_type == 'sicherstellung' else colors.HexColor('#f8fafc')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(order_table)
    
    # ===== FAHRZEUGDATEN =====
    story.append(Paragraph("Fahrzeugdaten", heading_style))
    vehicle_data = [
        [Paragraph("<b>Kennzeichen</b>", cell_style), Paragraph(job['license_plate'], cell_style)],
        [Paragraph("<b>FIN</b>", cell_style), Paragraph(job.get('vin') or '-', cell_style)],
        [Paragraph("<b>Abschleppgrund</b>", cell_style), Paragraph(wrap_text(job['tow_reason'], 50), cell_style)],
    ]
    # Zeige nur Behörde, NICHT die Dienstnummer des Mitarbeiters
    if job.get('created_by_authority'):
        vehicle_data.append([
            Paragraph("<b>Erfasst von</b>", cell_style), 
            Paragraph(f"{job.get('created_by_authority') or '-'}", cell_style)
        ])
    
    vehicle_table = Table(vehicle_data, colWidths=[4.5*cm, 12.5*cm])
    vehicle_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(vehicle_table)
    
    # ===== FUNDORT =====
    story.append(Paragraph("Fundort", heading_style))
    # Clean up address if it contains raw coordinates at the beginning
    display_address = job['location_address']
    if display_address and re.match(r'^\s*[-+]?\d{1,2}\.\d+,\s*[-+]?\d{1,3}\.\d+\s*[-|:]\s*', display_address):
        # Extract everything after the coordinates
        parts = re.split(r'^\s*[-+]?\d{1,2}\.\d+,\s*[-+]?\d{1,3}\.\d+\s*[-|:]\s*', display_address, 1)
        if len(parts) > 1 and parts[1].strip():
            display_address = parts[1].strip()
            
    # Also handle the exact case the user showed: "52.379462, 9.724245" where address is just coords
    if display_address and re.match(r'^\s*[-+]?\d{1,2}\.\d+,\s*[-+]?\d{1,3}\.\d+\s*$', display_address):
        display_address = "Keine genaue Adresse verfügbar (nur Koordinaten)"

    location_data = [
        [Paragraph("<b>Adresse</b>", cell_style), Paragraph(wrap_text(display_address, 55), cell_style)],
        [Paragraph("<b>Koordinaten</b>", cell_style), Paragraph(f"{job['location_lat']:.6f}, {job['location_lng']:.6f}", cell_style)],
    ]
    location_table = Table(location_data, colWidths=[4.5*cm, 12.5*cm])
    location_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(location_table)
    
    # ===== ZEITERFASSUNG =====
    story.append(Paragraph("Zeiterfassung", heading_style))
    timeline_data = [
        [
            Paragraph("<b>Schritt</b>", cell_style), 
            Paragraph("<b>Datum & Uhrzeit</b>", cell_style),
            Paragraph("<b>Status</b>", cell_style)
        ],
        [
            Paragraph("1. Meldung erfasst", cell_style), 
            Paragraph(format_datetime(job.get('created_at')), cell_style),
            Paragraph("✓" if job.get('created_at') else "-", cell_style)
        ],
        [
            Paragraph("2. Vor Ort angekommen", cell_style), 
            Paragraph(format_datetime(job.get('on_site_at')), cell_style),
            Paragraph("✓" if job.get('on_site_at') else "-", cell_style)
        ]
    ]

    if job.get('is_empty_trip'):
        timeline_data.append([
            Paragraph("3. Leerfahrt verzeichnet", cell_style),
            Paragraph(format_datetime(job.get('updated_at')), cell_style),
            Paragraph("✓" if job.get('status') in ['empty_trip', 'released'] else "-", cell_style)
        ])
    else:
        timeline_data.extend([
            [
                Paragraph("3. Abgeschleppt", cell_style), 
                Paragraph(format_datetime(job.get('towed_at')), cell_style),
                Paragraph("✓" if job.get('towed_at') else "-", cell_style)
            ],
            [
                Paragraph("4. Im Hof eingetroffen", cell_style), 
                Paragraph(format_datetime(job.get('in_yard_at')), cell_style),
                Paragraph("✓" if job.get('in_yard_at') else "-", cell_style)
            ],
            [
                Paragraph("5. Fahrzeug abgeholt", cell_style), 
                Paragraph(format_datetime(job.get('released_at')), cell_style),
                Paragraph("✓" if job.get('released_at') else "-", cell_style)
            ]
        ])
    timeline_table = Table(timeline_data, colWidths=[6*cm, 8*cm, 3*cm])
    timeline_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f97316')),  # Orange header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    story.append(timeline_table)
    
    # ===== HALTERDATEN (if released) =====
    if job.get('owner_first_name'):
        story.append(Paragraph("Halterdaten", heading_style))
        owner_data = [
            [Paragraph("<b>Name</b>", cell_style), Paragraph(f"{job.get('owner_first_name') or ''} {job.get('owner_last_name') or ''}", cell_style)],
            [Paragraph("<b>Adresse</b>", cell_style), Paragraph(wrap_text(job.get('owner_address') or '-', 55), cell_style)],
        ]
        owner_table = Table(owner_data, colWidths=[4.5*cm, 12.5*cm])
        owner_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ]))
        story.append(owner_table)
    
    # ===== ZAHLUNGSINFORMATIONEN =====
    has_payment = job.get('payment_method') and not (job.get('is_empty_trip') and job.get('payment_amount', 0) == 0)
    if has_payment:
        story.append(Paragraph("Zahlungsinformationen", heading_style))
        payment_method_text = "Bar" if job['payment_method'] == 'cash' else "Kartenzahlung"
        payment_data = [
            [Paragraph("<b>Zahlungsart</b>", cell_style), Paragraph(payment_method_text, cell_style)],
            [Paragraph("<b>Betrag</b>", cell_style), Paragraph(f"{job.get('payment_amount', 0):.2f} €", cell_style)],
        ]
        payment_table = Table(payment_data, colWidths=[4.5*cm, 12.5*cm])
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
            ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#dcfce7')),  # Green background for amount
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ]))
        story.append(payment_table)
    
    # ===== ABSCHLEPPDIENST =====
    if service:
        story.append(Paragraph("Abschleppdienst", heading_style))
        service_data = [
            [Paragraph("<b>Unternehmen</b>", cell_style), Paragraph(service.get('company_name') or '-', cell_style)],
            [Paragraph("<b>Telefon</b>", cell_style), Paragraph(service.get('phone') or '-', cell_style)],
            [Paragraph("<b>Hof-Adresse</b>", cell_style), Paragraph(wrap_text(service.get('yard_address') or '-', 55), cell_style)],
            [Paragraph("<b>Öffnungszeiten</b>", cell_style), Paragraph(service.get('opening_hours') or '-', cell_style)],
        ]
        service_table = Table(service_data, colWidths=[4.5*cm, 12.5*cm])
        service_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ]))
        story.append(service_table)
    
    # ===== BEMERKUNGEN =====
    if job.get('notes') or job.get('service_notes'):
        story.append(Paragraph("Bemerkungen", heading_style))
        notes_data = []
        if job.get('notes'):
            notes_data.append([
                Paragraph("<b>Behörde</b>", cell_style), 
                Paragraph(wrap_text(job['notes'], 55), cell_style)
            ])
        if job.get('service_notes'):
            notes_data.append([
                Paragraph("<b>Abschleppdienst</b>", cell_style), 
                Paragraph(wrap_text(job['service_notes'], 55), cell_style)
            ])
        notes_table = Table(notes_data, colWidths=[4.5*cm, 12.5*cm])
        notes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ]))
        story.append(notes_table)
    
    # ===== FOOTER =====
    story.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#94a3b8'),
        alignment=1
    )
    story.append(Paragraph(f"Erstellt am: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} UTC", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Abschleppprotokoll_{job['job_number']}.pdf"}
    )

# ==================== EXPORT ENDPOINTS ====================

@api_router.get("/export/jobs/csv")
async def export_jobs_csv(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Export jobs as CSV"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Build query
    query = {}
    if status:
        query["status"] = status
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_query
    
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Create CSV
    output = BytesIO()
    output.write('\ufeff'.encode('utf-8'))  # UTF-8 BOM for Excel
    
    import csv
    import io
    text_output = io.StringIO()
    
    fieldnames = [
        'Auftragsnummer', 'Kennzeichen', 'FIN', 'Abschleppgrund', 'Status',
        'Standort', 'Behörde', 'Abschleppdienst',
        'Erstellt am', 'Vor Ort', 'Abgeschleppt', 'Im Hof', 'Abgeholt',
        'Halter Name', 'Halter Adresse', 'Zahlungsart', 'Betrag'
    ]
    
    writer = csv.DictWriter(text_output, fieldnames=fieldnames, delimiter=';')
    writer.writeheader()
    
    status_map = {
        'pending': 'Ausstehend', 'assigned': 'Zugewiesen', 'on_site': 'Vor Ort',
        'towed': 'Abgeschleppt', 'in_yard': 'Im Hof', 'released': 'Abgeholt'
    }
    
    for job in jobs:
        writer.writerow({
            'Auftragsnummer': job.get('job_number', ''),
            'Kennzeichen': job.get('license_plate', ''),
            'FIN': job.get('vin', ''),
            'Abschleppgrund': job.get('tow_reason', ''),
            'Status': status_map.get(job.get('status', ''), job.get('status', '')),
            'Standort': job.get('location_address', ''),
            'Behörde': job.get('created_by_authority', ''),
            'Abschleppdienst': job.get('assigned_service_name', ''),
            'Erstellt am': job.get('created_at', '')[:19].replace('T', ' ') if job.get('created_at') else '',
            'Vor Ort': job.get('on_site_at', '')[:19].replace('T', ' ') if job.get('on_site_at') else '',
            'Abgeschleppt': job.get('towed_at', '')[:19].replace('T', ' ') if job.get('towed_at') else '',
            'Im Hof': job.get('in_yard_at', '')[:19].replace('T', ' ') if job.get('in_yard_at') else '',
            'Abgeholt': job.get('released_at', '')[:19].replace('T', ' ') if job.get('released_at') else '',
            'Halter Name': f"{job.get('owner_first_name', '')} {job.get('owner_last_name', '')}".strip(),
            'Halter Adresse': job.get('owner_address', ''),
            'Zahlungsart': 'Bar' if job.get('payment_method') == 'cash' else ('Karte' if job.get('payment_method') == 'card' else ''),
            'Betrag': f"{job.get('payment_amount', 0):.2f}" if job.get('payment_amount') else ''
        })
    
    await log_audit("EXPORT_CSV", user["id"], user["name"], {"count": len(jobs)})
    
    csv_content = text_output.getvalue().encode('utf-8-sig')
    
    return StreamingResponse(
        BytesIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=Auftraege_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )

@api_router.get("/export/jobs/excel")
async def export_jobs_excel(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Export jobs as Excel"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Build query
    query = {}
    if status:
        query["status"] = status
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_query
    
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aufträge"
    
    # Header styling
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        'Auftragsnummer', 'Kennzeichen', 'FIN', 'Abschleppgrund', 'Status',
        'Standort', 'Behörde', 'Abschleppdienst',
        'Erstellt am', 'Vor Ort', 'Abgeschleppt', 'Im Hof', 'Abgeholt',
        'Halter Name', 'Halter Adresse', 'Zahlungsart', 'Betrag'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    status_map = {
        'pending': 'Ausstehend', 'assigned': 'Zugewiesen', 'on_site': 'Vor Ort',
        'towed': 'Abgeschleppt', 'in_yard': 'Im Hof', 'released': 'Abgeholt'
    }
    
    for row, job in enumerate(jobs, 2):
        data = [
            job.get('job_number', ''),
            job.get('license_plate', ''),
            job.get('vin', ''),
            job.get('tow_reason', ''),
            status_map.get(job.get('status', ''), job.get('status', '')),
            job.get('location_address', ''),
            job.get('created_by_authority', ''),
            job.get('assigned_service_name', ''),
            job.get('created_at', '')[:19].replace('T', ' ') if job.get('created_at') else '',
            job.get('on_site_at', '')[:19].replace('T', ' ') if job.get('on_site_at') else '',
            job.get('towed_at', '')[:19].replace('T', ' ') if job.get('towed_at') else '',
            job.get('in_yard_at', '')[:19].replace('T', ' ') if job.get('in_yard_at') else '',
            job.get('released_at', '')[:19].replace('T', ' ') if job.get('released_at') else '',
            f"{job.get('owner_first_name', '')} {job.get('owner_last_name', '')}".strip(),
            job.get('owner_address', ''),
            'Bar' if job.get('payment_method') == 'cash' else ('Karte' if job.get('payment_method') == 'card' else ''),
            job.get('payment_amount', 0) if job.get('payment_amount') else ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    await log_audit("EXPORT_EXCEL", user["id"], user["name"], {"count": len(jobs)})
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Auftraege_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

# ==================== AUDIT LOG ENDPOINTS ====================

class AuditLogResponse(BaseModel):
    id: str
    timestamp: str
    action: str
    user_id: Optional[str] = "system"
    user_name: Optional[str] = "System"
    details: Dict[str, Any] = {}

@api_router.get("/admin/audit-logs", response_model=List[AuditLogResponse])
async def get_audit_logs(
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    user: dict = Depends(get_current_user)
):
    """Get audit logs (Admin only)"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    query = {}
    if action:
        query["action"] = {"$regex": action, "$options": "i"}
    if user_id:
        query["user_id"] = user_id
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        query["timestamp"] = date_query
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    
    return [AuditLogResponse(**log) for log in logs]

@api_router.get("/admin/audit-logs/count")
async def get_audit_logs_count(user: dict = Depends(get_current_user)):
    """Get total audit log count"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    count = await db.audit_logs.count_documents({})
    return {"count": count}

# ==================== FULL-TEXT SEARCH ====================

@api_router.get("/search/jobs")
async def search_jobs_fulltext(
    q: str,
    limit: int = 50,
    skip: int = 0,
    user: dict = Depends(get_current_user)
):
    """Full-text search across all job fields"""
    if not q or len(q) < 2:
        raise HTTPException(status_code=400, detail="Suchbegriff muss mindestens 2 Zeichen haben")
    
    # Build search query
    search_regex = {"$regex": q, "$options": "i"}
    
    query = {
        "$or": [
            {"job_number": search_regex},
            {"license_plate": search_regex},
            {"vin": search_regex},
            {"tow_reason": search_regex},
            {"location_address": search_regex},
            {"created_by_name": search_regex},
            {"created_by_authority": search_regex},
            {"created_by_dienstnummer": search_regex},
            {"assigned_service_name": search_regex},
            {"owner_first_name": search_regex},
            {"owner_last_name": search_regex},
            {"owner_address": search_regex},
            {"notes": search_regex},
            {"service_notes": search_regex}
        ]
    }
    
    # Apply role-based filtering
    if user["role"] == UserRole.AUTHORITY:
        if user.get("is_main_authority"):
            query["authority_id"] = user["id"]
        else:
            query["created_by_id"] = user["id"]
    elif user["role"] == UserRole.TOWING_SERVICE:
        query["assigned_service_id"] = user["id"]
    
    total = await db.jobs.count_documents(query)
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "total": total,
        "results": [JobResponse(**j) for j in jobs],
        "query": q
    }

# ==================== PAGINATION FOR JOBS ====================

@api_router.get("/jobs/paginated")
async def get_jobs_paginated(
    page: int = 1,
    per_page: int = 20,
    status: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get paginated jobs with filters"""
    query = {}
    
    # Role-based filtering
    if user["role"] == UserRole.AUTHORITY:
        if user.get("is_main_authority"):
            query["authority_id"] = user["id"]
        else:
            query["created_by_id"] = user["id"]
    elif user["role"] == UserRole.TOWING_SERVICE:
        query["assigned_service_id"] = user["id"]
    
    # Apply filters
    if status:
        query["status"] = status
    
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_query
    
    if search:
        search_upper = search.upper()
        query["$or"] = [
            {"license_plate": {"$regex": search_upper, "$options": "i"}},
            {"vin": {"$regex": search_upper, "$options": "i"}},
            {"job_number": {"$regex": search_upper, "$options": "i"}}
        ]
    
    # Calculate pagination
    skip = (page - 1) * per_page
    total = await db.jobs.count_documents(query)
    total_pages = math.ceil(total / per_page) if total > 0 else 1
    
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "data": [JobResponse(**j) for j in jobs],
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1
        }
    }

# ==================== DATABASE BACKUP ====================

@api_router.post("/admin/backup")
async def create_backup(
    backup_type: str = "database",
    user: dict = Depends(get_current_user)
):
    """Create a backup using the full backup service (Admin only)"""
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin only")
    
    try:
        if backup_type == "database":
            result = await backup_service.create_database_backup(
                triggered_by_user_id=user["id"],
                retention_class="daily"
            )
        elif backup_type == "storage":
            result = await backup_service.create_storage_backup(
                triggered_by_user_id=user["id"],
                retention_class="daily"
            )
        else:
            raise HTTPException(status_code=400, detail="Ungültiger Backup-Typ. Erlaubt: database, storage")
        
        if result.get("status") == "success":
            return {
                "message": "Backup erfolgreich erstellt",
                "filename": result.get("filename"),
                "size": result.get("size_bytes"),
                "supabase_uploaded": result.get("supabase_uploaded", False),
                "supabase_path": result.get("supabase_path")
            }
        else:
            raise HTTPException(status_code=500, detail=result.get("error", "Backup fehlgeschlagen"))
            
    except Exception as e:
        logger.error(f"Backup failed: {e}")
        raise HTTPException(status_code=500, detail=f"Backup fehlgeschlagen: {str(e)}")

# OLD BACKUP ROUTES REMOVED - Using new BackupService instead

# ==================== FILE UPLOAD ====================

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    file_id = str(uuid.uuid4())
    file_ext = Path(file.filename).suffix if file.filename else ".jpg"
    file_path = UPLOAD_DIR / f"{file_id}{file_ext}"
    
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    return {"file_id": file_id, "filename": f"{file_id}{file_ext}"}

@api_router.get("/uploads/{filename}")
async def get_upload(filename: str):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path)

# ==================== ROOT ====================

@api_router.get("/")
async def root():
    return {"message": "Abschlepp-Management API", "version": "1.0.0"}

# ==================== DSGVO DATA CLEANUP & STEUERRECHTLICHE AUFBEWAHRUNG ====================

"""
DATENSCHUTZ- UND AUFBEWAHRUNGSKONZEPT:

1. DSGVO (Datenschutz-Grundverordnung):
   - Personenbezogene Daten werden nach 6 Monaten (180 Tage) anonymisiert
   - Betrifft: Kennzeichen, FIN, Halterdaten, Fotos
   
2. Steuerrecht (§ 147 AO / § 257 HGB):
   - Rechnungen und Buchungsbelege: 10 Jahre Aufbewahrungspflicht
   - Betrifft: Rechnungsnummer, Beträge, Zahlungsdaten, Auftraggeber (Behörde), Dienstleister
   
3. Ablauf:
   - Nach 6 Monaten: Personenbezogene Daten anonymisiert, Rechnungsdaten bleiben
   - Nach 10 Jahren: Gesamter Datensatz kann gelöscht werden
"""

# Felder die ANONYMISIERT werden (personenbezogen)
PERSONAL_DATA_FIELDS = [
    "license_plate",      # Kennzeichen
    "vin",                # Fahrzeug-Identifizierungsnummer
    "owner_name",         # Haltername
    "owner_address",      # Halteradresse
    "owner_phone",        # Haltertelefon
    "driver_name",        # Fahrername (bei Leerfahrt)
    "driver_address",     # Fahreradresse
    "vehicle_brand",      # Fahrzeugmarke (kann identifizierend sein)
    "vehicle_model",      # Fahrzeugmodell
    "vehicle_color",      # Fahrzeugfarbe
    "location_address",   # Abschleppstandort (kann sensibel sein)
]

# Felder die BEHALTEN werden (für Rechnungen/Steuer)
INVOICE_DATA_FIELDS = [
    "id",                 # Interne ID
    "job_number",         # Rechnungsnummer
    "created_at",         # Erstellungsdatum
    "released_at",        # Freigabedatum
    "tow_cost",           # Abschleppkosten
    "daily_cost",         # Standgebühren
    "total_cost",         # Gesamtkosten
    "processing_fee",     # Bearbeitungsgebühr
    "payment_method",     # Zahlungsart
    "payment_received",   # Zahlung erhalten
    "paid_amount",        # Bezahlter Betrag
    "authority_id",       # Auftraggeber-ID
    "authority_name",     # Auftraggeber (für Rechnung)
    "assigned_service_id", # Dienstleister-ID
    "service_name",       # Dienstleister (für Rechnung)
    "tow_reason",         # Abschleppgrund (für Buchungszweck)
    "status",             # Status
    "is_sicherstellung",  # Sicherstellung-Flag
    "days_in_yard",       # Standtage
]


async def dsgvo_data_cleanup():
    """
    DSGVO-konformer Cronjob zur automatischen Anonymisierung von personenbezogenen Daten.
    
    WICHTIG: Unterscheidet zwischen:
    - Personenbezogene Daten → nach 6 Monaten anonymisieren
    - Rechnungsdaten → 10 Jahre aufbewahren (Steuerrecht)
    
    Läuft täglich um 03:00 Uhr.
    """
    logger.info(f"DSGVO Data Cleanup gestartet - Personendaten-Retention: {DSGVO_RETENTION_DAYS} Tage, Rechnungs-Retention: {INVOICE_RETENTION_YEARS} Jahre")
    
    try:
        cutoff_date_personal = datetime.now(timezone.utc) - timedelta(days=DSGVO_RETENTION_DAYS)
        cutoff_date_str = cutoff_date_personal.isoformat()
        
        # Find all released jobs older than personal data retention period
        # that haven't been anonymized yet
        query = {
            "status": "released",
            "released_at": {"$lt": cutoff_date_str},
            "personal_data_anonymized": {"$ne": True}  # New flag to track personal data anonymization
        }
        
        jobs_to_anonymize = await db.jobs.find(query).to_list(length=1000)
        
        if not jobs_to_anonymize:
            logger.info("DSGVO Cleanup: Keine Aufträge zur Anonymisierung gefunden.")
            # Still run invoice cleanup
            await invoice_data_cleanup()
            return
        
        anonymized_count = 0
        
        for job in jobs_to_anonymize:
            job_id = job.get("id") or str(job.get("_id"))
            original_plate = job.get("license_plate", "")
            
            # Anonymize ONLY personal data fields, keep invoice data
            anonymization_data = {
                "personal_data_anonymized": True,
                "personal_data_anonymized_at": datetime.now(timezone.utc).isoformat(),
                "original_license_plate_hash": str(hash(original_plate)),  # Keep hash for audit trail
                "photos": [],  # Clear photos as they may contain personal data
                # Legacy flag for backwards compatibility
                "anonymized": True,
                "anonymized_at": datetime.now(timezone.utc).isoformat(),
            }
            
            # Anonymize each personal data field that exists
            for field in PERSONAL_DATA_FIELDS:
                if job.get(field):
                    anonymization_data[field] = "*** (DSGVO-Anonymisiert)"
            
            await db.jobs.update_one(
                {"id": job_id} if job.get("id") else {"_id": job["_id"]},
                {"$set": anonymization_data}
            )
            
            anonymized_count += 1
        
        # Log the cleanup action to audit log
        await log_audit("DSGVO_PERSONAL_DATA_CLEANUP", "SYSTEM", "Automatischer DSGVO Cronjob", {
            "jobs_anonymized": anonymized_count,
            "personal_data_retention_days": DSGVO_RETENTION_DAYS,
            "invoice_retention_years": INVOICE_RETENTION_YEARS,
            "cutoff_date": cutoff_date_str,
            "note": "Personenbezogene Daten anonymisiert, Rechnungsdaten bleiben erhalten (§ 147 AO)"
        })
        
        logger.info(f"DSGVO Cleanup abgeschlossen: {anonymized_count} Aufträge - Personendaten anonymisiert, Rechnungsdaten erhalten.")
        
        # Also run invoice cleanup for very old records
        await invoice_data_cleanup()
        
    except Exception as e:
        logger.error(f"DSGVO Cleanup Fehler: {e}", exc_info=True)
        await log_audit("DSGVO_CLEANUP_ERROR", "SYSTEM", "Automatischer DSGVO Cronjob", {
            "error": str(e)
        })


async def invoice_data_cleanup():
    """
    Steuerrechtlicher Cleanup: Löscht vollständige Datensätze nach 10 Jahren.
    
    Nach Ablauf der steuerrechtlichen Aufbewahrungsfrist (10 Jahre) können
    auch die Rechnungsdaten gelöscht werden.
    """
    try:
        cutoff_date_invoice = datetime.now(timezone.utc) - timedelta(days=INVOICE_RETENTION_YEARS * 365)
        cutoff_date_str = cutoff_date_invoice.isoformat()
        
        # Find all jobs older than invoice retention period
        query = {
            "released_at": {"$lt": cutoff_date_str},
            "invoice_data_deleted": {"$ne": True}
        }
        
        jobs_to_delete = await db.jobs.find(query).to_list(length=1000)
        
        if not jobs_to_delete:
            logger.info("Steuerrecht Cleanup: Keine Aufträge zur Löschung gefunden (alle jünger als 10 Jahre).")
            return
        
        deleted_count = 0
        archived_job_numbers = []
        
        for job in jobs_to_delete:
            job_id = job.get("id") or str(job.get("_id"))
            job_number = job.get("job_number", "Unbekannt")
            archived_job_numbers.append(job_number)
            
            # Option 1: Vollständig löschen
            # await db.jobs.delete_one({"id": job_id} if job.get("id") else {"_id": job["_id"]})
            
            # Option 2: Markieren als gelöscht (empfohlen für Audit-Trail)
            await db.jobs.update_one(
                {"id": job_id} if job.get("id") else {"_id": job["_id"]},
                {"$set": {
                    "invoice_data_deleted": True,
                    "invoice_data_deleted_at": datetime.now(timezone.utc).isoformat(),
                    "deletion_reason": f"Steuerrechtliche Aufbewahrungsfrist ({INVOICE_RETENTION_YEARS} Jahre) abgelaufen"
                }}
            )
            
            deleted_count += 1
        
        if deleted_count > 0:
            await log_audit("STEUERRECHT_DATA_CLEANUP", "SYSTEM", "Automatischer Steuerrecht Cronjob", {
                "jobs_marked_deleted": deleted_count,
                "retention_years": INVOICE_RETENTION_YEARS,
                "cutoff_date": cutoff_date_str,
                "job_numbers": archived_job_numbers[:50],  # Log first 50 for reference
                "note": "Aufbewahrungsfrist nach § 147 AO / § 257 HGB abgelaufen"
            })
            
            logger.info(f"Steuerrecht Cleanup: {deleted_count} Aufträge zur Löschung markiert (älter als {INVOICE_RETENTION_YEARS} Jahre).")
        
    except Exception as e:
        logger.error(f"Steuerrecht Cleanup Fehler: {e}", exc_info=True)


# Admin endpoint to manually trigger DSGVO cleanup (for testing)
@api_router.post("/admin/trigger-cleanup")
async def trigger_dsgvo_cleanup(user: dict = Depends(get_current_user)):
    """Manueller Trigger für den DSGVO Cleanup Job (nur Admin)"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Nur Administratoren können den Cleanup manuell starten")
    
    await dsgvo_data_cleanup()
    
    return {
        "message": "DSGVO & Steuerrecht Cleanup wurde gestartet",
        "personal_data_retention_days": DSGVO_RETENTION_DAYS,
        "invoice_retention_years": INVOICE_RETENTION_YEARS,
        "note": "Personendaten werden nach 6 Monaten anonymisiert, Rechnungsdaten nach 10 Jahren gelöscht"
    }


# Admin endpoint to get DSGVO status/info
@api_router.get("/admin/dsgvo-status")
async def get_dsgvo_status(user: dict = Depends(get_current_user)):
    """Zeigt den Status der DSGVO- und steuerrechtlichen Datenhaltung"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Nur für Administratoren")
    
    now = datetime.now(timezone.utc)
    cutoff_date_personal = now - timedelta(days=DSGVO_RETENTION_DAYS)
    cutoff_date_invoice = now - timedelta(days=INVOICE_RETENTION_YEARS * 365)
    
    # Count jobs pending personal data anonymization (older than 6 months)
    pending_personal_anonymization = await db.jobs.count_documents({
        "status": "released",
        "released_at": {"$lt": cutoff_date_personal.isoformat()},
        "personal_data_anonymized": {"$ne": True}
    })
    
    # Count already anonymized jobs (personal data)
    personal_data_anonymized = await db.jobs.count_documents({
        "personal_data_anonymized": True
    })
    
    # Count jobs pending invoice deletion (older than 10 years)
    pending_invoice_deletion = await db.jobs.count_documents({
        "released_at": {"$lt": cutoff_date_invoice.isoformat()},
        "invoice_data_deleted": {"$ne": True}
    })
    
    # Count deleted invoice data
    invoice_data_deleted = await db.jobs.count_documents({
        "invoice_data_deleted": True
    })
    
    # Get last cleanups from audit log
    last_personal_cleanup = await db.audit_logs.find_one(
        {"action": {"$in": ["DSGVO_PERSONAL_DATA_CLEANUP", "DSGVO_DATA_CLEANUP"]}},
        sort=[("timestamp", -1)]
    )
    
    last_invoice_cleanup = await db.audit_logs.find_one(
        {"action": "STEUERRECHT_DATA_CLEANUP"},
        sort=[("timestamp", -1)]
    )
    
    return {
        # DSGVO (Personal Data)
        "dsgvo": {
            "retention_days": DSGVO_RETENTION_DAYS,
            "retention_months": DSGVO_RETENTION_DAYS // 30,
            "cutoff_date": cutoff_date_personal.isoformat(),
            "pending_anonymization": pending_personal_anonymization,
            "already_anonymized": personal_data_anonymized,
            "last_cleanup": last_personal_cleanup["timestamp"] if last_personal_cleanup else None,
            "description": "Personenbezogene Daten (Kennzeichen, Halterdaten, Fotos)"
        },
        # Steuerrecht (Invoice Data)
        "steuerrecht": {
            "retention_years": INVOICE_RETENTION_YEARS,
            "retention_days": INVOICE_RETENTION_YEARS * 365,
            "cutoff_date": cutoff_date_invoice.isoformat(),
            "pending_deletion": pending_invoice_deletion,
            "already_deleted": invoice_data_deleted,
            "last_cleanup": last_invoice_cleanup["timestamp"] if last_invoice_cleanup else None,
            "legal_basis": "§ 147 AO / § 257 HGB",
            "description": "Rechnungsdaten (Beträge, Rechnungsnummer, Zahlungen)"
        },
        "scheduler_running": scheduler.running
    }

# ==================== BACKUP SERVICE ====================
from backup_service import BackupService

backup_service = BackupService(db, mongo_url, os.environ['DB_NAME'])

# ==================== BACKUP ROUTES (ADMIN ONLY) ====================

@api_router.get("/admin/backups/system-status")
async def get_backup_system_status(current_user: dict = Depends(get_current_user)):
    """System-Status der Backups - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.get_system_status()

@api_router.get("/admin/backups/cloud")
async def list_cloud_backups_route(current_user: dict = Depends(get_current_user)):
    """Liste aller Backups in Supabase Cloud - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    cloud_backups = await backup_service.list_cloud_backups()
    return cloud_backups

@api_router.post("/admin/backups/cloud/restore")
async def restore_from_cloud_route(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Backup direkt von Supabase Cloud wiederherstellen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups wiederherstellen")
    
    data = await request.json()
    cloud_path = data.get("cloud_path")
    confirm = data.get("confirm", False)
    
    if not cloud_path:
        raise HTTPException(status_code=400, detail="cloud_path ist erforderlich")
    
    result = await backup_service.restore_from_cloud(
        cloud_path=cloud_path,
        triggered_by_user_id=current_user.get("id"),
        confirm=confirm
    )
    
    return result

@api_router.get("/admin/backups/health")
async def get_backup_health_route(current_user: dict = Depends(get_current_user)):
    """Backup-Gesundheitsstatus - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.get_backup_health()

@api_router.post("/admin/backups/verify-all")
async def verify_all_backups_route(current_user: dict = Depends(get_current_user)):
    """Alle Backups verifizieren - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.verify_all_backups()

@api_router.post("/admin/backups/{backup_id}/verify")
async def verify_backup_route(
    backup_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Einzelnes Backup verifizieren - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.verify_backup(backup_id)

@api_router.get("/admin/backups/schedule")
async def get_backup_schedule_route(current_user: dict = Depends(get_current_user)):
    """Backup-Zeitplan-Einstellungen abrufen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.get_schedule_settings()

@api_router.put("/admin/backups/schedule")
async def update_backup_schedule_route(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Backup-Zeitplan-Einstellungen aktualisieren - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    data = await request.json()
    return await backup_service.update_schedule_settings(data, current_user.get("id"))

@api_router.get("/admin/backups/storage-stats")
async def get_storage_stats_route(current_user: dict = Depends(get_current_user)):
    """Speicherplatz-Statistiken - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.get_storage_stats()

@api_router.post("/admin/backups/cleanup")
async def cleanup_backups_route(current_user: dict = Depends(get_current_user)):
    """Alte Backups bereinigen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.cleanup_old_backups()

@api_router.post("/admin/backups/emergency-cleanup")
async def emergency_cleanup_route(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Notfall-Bereinigung - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    data = await request.json()
    target_mb = data.get("target_mb", 500)
    return await backup_service.emergency_cleanup(target_mb)

@api_router.get("/admin/backups/encryption")
async def get_encryption_settings_route(current_user: dict = Depends(get_current_user)):
    """Verschlüsselungseinstellungen abrufen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.get_encryption_settings()

@api_router.post("/admin/backups/send-test-email")
async def send_test_email_route(current_user: dict = Depends(get_current_user)):
    """Test-E-Mail senden - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können E-Mails senden")
    
    try:
        result = await email_service.send_backup_failure_alert(
            backup_type="Test-Backup",
            error_message="Dies ist eine Test-E-Mail um die Backup-Benachrichtigungen zu prüfen. Keine Aktion erforderlich.",
            backup_id="TEST-" + datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
        )
        return {
            "success": True,
            "message": f"Test-E-Mail wurde an {ADMIN_EMAIL} gesendet",
            "message_id": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"E-Mail-Versand fehlgeschlagen: {str(e)}")

@api_router.post("/admin/backups/send-weekly-report")
async def send_weekly_report_manual_route(current_user: dict = Depends(get_current_user)):
    """Wöchentlichen Backup-Report manuell senden - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Reports senden")
    
    try:
        await scheduled_weekly_backup_report()
        return {
            "success": True,
            "message": f"Wöchentlicher Report wurde an {ADMIN_EMAIL} gesendet"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Report-Versand fehlgeschlagen: {str(e)}")

@api_router.put("/admin/backups/encryption")
async def update_encryption_settings_route(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Verschlüsselung aktivieren/deaktivieren - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    data = await request.json()
    enabled = data.get("enabled", False)
    return await backup_service.update_encryption_settings(enabled, current_user.get("id"))

@api_router.get("/admin/backups")
async def list_backups_route(
    backup_type: Optional[str] = None,
    retention_class: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """Liste aller Backups - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    backups = await backup_service.list_backups(backup_type, retention_class, status, limit)
    return backups

@api_router.post("/admin/backups/run-database-backup")
async def run_database_backup_route(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Manuelles Datenbank-Backup starten - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups erstellen")
    
    result = await backup_service.create_database_backup(
        triggered_by_user_id=current_user.get("id"),
        retention_class="daily"
    )
    return result

@api_router.post("/admin/backups/run-storage-backup")
async def run_storage_backup_route(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Manuelles Storage-Backup starten - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups erstellen")
    
    result = await backup_service.create_storage_backup(
        triggered_by_user_id=current_user.get("id"),
        retention_class="daily"
    )
    return result

@api_router.post("/admin/backups/run-full-backup")
async def run_full_backup_route(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Manuelles Komplett-Backup (DB + Storage) starten - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups erstellen")
    
    result = await backup_service.create_full_backup(
        triggered_by_user_id=current_user.get("id"),
        retention_class="daily"
    )
    return result

@api_router.get("/admin/backups/{backup_id}/download")
async def download_backup_route(
    backup_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Backup herunterladen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups herunterladen")
    
    backup_path = await backup_service.get_backup_file_path(backup_id)
    if not backup_path:
        raise HTTPException(status_code=404, detail="Backup-Datei nicht gefunden")
    
    return FileResponse(
        path=str(backup_path),
        filename=backup_path.name,
        media_type="application/octet-stream"
    )

@api_router.post("/admin/backups/{backup_id}/restore-database")
async def restore_database_backup_route(
    backup_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Datenbank aus Backup wiederherstellen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups wiederherstellen")
    
    confirm = data.get("confirm", False)
    result = await backup_service.restore_database_backup(
        backup_id=backup_id,
        triggered_by_user_id=current_user.get("id"),
        confirm=confirm
    )
    return result

@api_router.post("/admin/backups/{backup_id}/restore-storage")
async def restore_storage_backup_route(
    backup_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Storage aus Backup wiederherstellen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups wiederherstellen")
    
    confirm = data.get("confirm", False)
    result = await backup_service.restore_storage_backup(
        backup_id=backup_id,
        triggered_by_user_id=current_user.get("id"),
        confirm=confirm
    )
    return result

@api_router.delete("/admin/backups/{backup_id}")
async def delete_backup_route(
    backup_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Backup löschen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups löschen")
    
    result = await backup_service.delete_backup(
        backup_id=backup_id,
        triggered_by_user_id=current_user.get("id")
    )
    return result

@api_router.get("/admin/backups/{backup_id}")
async def get_backup_details_route(
    backup_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Details eines Backups - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    backup = await backup_service.get_backup(backup_id)
    if not backup:
        raise HTTPException(status_code=404, detail="Backup nicht gefunden")
    return backup

# Include router
app.include_router(api_router)

_allowed_origins = os.environ.get('CORS_ORIGINS', 'http://localhost:3000').split(',')
if "http://localhost:3000" not in _allowed_origins:
    _allowed_origins.append("http://localhost:3000")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=_allowed_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== BACKUP SERVICE ====================
from backup_service import BackupService

backup_service = BackupService(db, mongo_url, os.environ['DB_NAME'])

# ==================== BACKUP ROUTES (ADMIN ONLY) ====================

@api_router.get("/admin/backups")
async def list_backups(
    backup_type: Optional[str] = None,
    retention_class: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """Liste aller Backups - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    backups = await backup_service.list_backups(backup_type, retention_class, status, limit)
    return backups

@api_router.get("/admin/backups/system-status")
async def get_backup_system_status(current_user: dict = Depends(get_current_user)):
    """System-Status der Backups - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    return await backup_service.get_system_status()

@api_router.post("/admin/backups/run-database-backup")
async def run_database_backup(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Manuelles Datenbank-Backup starten - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups erstellen")
    
    result = await backup_service.create_database_backup(
        triggered_by_user_id=current_user.get("id"),
        retention_class="daily"
    )
    return result

@api_router.post("/admin/backups/run-storage-backup")
async def run_storage_backup(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Manuelles Storage-Backup starten - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups erstellen")
    
    result = await backup_service.create_storage_backup(
        triggered_by_user_id=current_user.get("id"),
        retention_class="daily"
    )
    return result

@api_router.post("/admin/backups/run-full-backup")
async def run_full_backup(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Manuelles Komplett-Backup (DB + Storage) starten - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups erstellen")
    
    result = await backup_service.create_full_backup(
        triggered_by_user_id=current_user.get("id"),
        retention_class="daily"
    )
    return result

@api_router.get("/admin/backups/{backup_id}")
async def get_backup_details(
    backup_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Details eines Backups - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups verwalten")
    
    backup = await backup_service.get_backup(backup_id)
    if not backup:
        raise HTTPException(status_code=404, detail="Backup nicht gefunden")
    return backup

@api_router.get("/admin/backups/{backup_id}/download")
async def download_backup(
    backup_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Backup herunterladen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups herunterladen")
    
    backup_path = await backup_service.get_backup_file_path(backup_id)
    if not backup_path:
        raise HTTPException(status_code=404, detail="Backup-Datei nicht gefunden")
    
    return FileResponse(
        path=str(backup_path),
        filename=backup_path.name,
        media_type="application/octet-stream"
    )

@api_router.post("/admin/backups/{backup_id}/restore-database")
async def restore_database_backup(
    backup_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Datenbank aus Backup wiederherstellen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups wiederherstellen")
    
    confirm = data.get("confirm", False)
    result = await backup_service.restore_database_backup(
        backup_id=backup_id,
        triggered_by_user_id=current_user.get("id"),
        confirm=confirm
    )
    return result

@api_router.post("/admin/backups/{backup_id}/restore-storage")
async def restore_storage_backup(
    backup_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Storage aus Backup wiederherstellen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups wiederherstellen")
    
    confirm = data.get("confirm", False)
    result = await backup_service.restore_storage_backup(
        backup_id=backup_id,
        triggered_by_user_id=current_user.get("id"),
        confirm=confirm
    )
    return result

@api_router.delete("/admin/backups/{backup_id}")
async def delete_backup(
    backup_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Backup löschen - nur Admin"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Nur Admins können Backups löschen")
    
    result = await backup_service.delete_backup(
        backup_id=backup_id,
        triggered_by_user_id=current_user.get("id")
    )
    return result

# ==================== SCHEDULED BACKUP FUNCTIONS ====================

async def scheduled_database_backup():
    """Geplantes tägliches Datenbank-Backup"""
    logger.info("Starting scheduled database backup")
    try:
        result = await backup_service.create_database_backup(
            triggered_by_user_id=None,
            retention_class="daily"
        )
        logger.info(f"Scheduled database backup completed: {result}")
    except Exception as e:
        logger.error(f"Scheduled database backup failed: {e}")

async def scheduled_storage_backup():
    """Geplantes tägliches Storage-Backup"""
    logger.info("Starting scheduled storage backup")
    try:
        result = await backup_service.create_storage_backup(
            triggered_by_user_id=None,
            retention_class="daily"
        )
        logger.info(f"Scheduled storage backup completed: {result}")
    except Exception as e:
        logger.error(f"Scheduled storage backup failed: {e}")

async def scheduled_monthly_backup():
    """Geplantes monatliches Komplett-Backup"""
    logger.info("Starting scheduled monthly backup")
    try:
        result = await backup_service.create_full_backup(
            triggered_by_user_id=None,
            retention_class="monthly"
        )
        logger.info(f"Scheduled monthly backup completed: {result}")
    except Exception as e:
        logger.error(f"Scheduled monthly backup failed: {e}")

async def scheduled_retention_cleanup():
    """Geplante Aufräumung alter Backups"""
    logger.info("Starting retention cleanup")
    try:
        result = await backup_service.apply_retention_rules()
        logger.info(f"Retention cleanup completed: {result}")
    except Exception as e:
        logger.error(f"Retention cleanup failed: {e}")

async def scheduled_weekly_backup_report():
    """Sendet den wöchentlichen Backup-Status-Report per E-Mail"""
    logger.info("Starting weekly backup report generation")
    try:
        # Berechne Statistiken der letzten 7 Tage
        week_ago = datetime.now(timezone.utc) - timedelta(days=7)
        
        # Alle Backups der letzten Woche
        all_backups = await db.backup_jobs.find({
            "created_at": {"$gte": week_ago.isoformat()}
        }).to_list(1000)
        
        total_backups = len(all_backups)
        failed_backups = sum(1 for b in all_backups if b.get("status") == "failed")
        db_backups = sum(1 for b in all_backups if b.get("backup_type") == "database")
        storage_backups = sum(1 for b in all_backups if b.get("backup_type") == "storage")
        cloud_uploads = sum(1 for b in all_backups if b.get("supabase_uploaded"))
        
        # Gesamtgröße
        total_size_bytes = sum(b.get("size_bytes", 0) for b in all_backups if b.get("size_bytes"))
        total_size_mb = total_size_bytes / (1024 * 1024) if total_size_bytes else 0
        
        # Letztes erfolgreiches Backup
        successful_backups = [b for b in all_backups if b.get("status") == "success"]
        last_successful = "Keines" if not successful_backups else max(
            successful_backups, key=lambda x: x.get("created_at", "")
        ).get("created_at", "")[:16].replace("T", " ")
        
        # Ältestes verfügbares Backup (aus allen, nicht nur letzte Woche)
        oldest_backup_doc = await db.backup_jobs.find_one(
            {"status": "success"},
            sort=[("created_at", 1)]
        )
        oldest_backup = oldest_backup_doc.get("created_at", "")[:10] if oldest_backup_doc else "N/A"
        
        stats = {
            "total_backups": total_backups,
            "failed_backups": failed_backups,
            "db_backups": db_backups,
            "storage_backups": storage_backups,
            "cloud_uploads": cloud_uploads,
            "total_size_mb": total_size_mb,
            "last_successful": last_successful,
            "oldest_backup": oldest_backup
        }
        
        await email_service.send_weekly_backup_report(stats)
        logger.info(f"Weekly backup report sent successfully: {stats}")
        
    except Exception as e:
        logger.error(f"Weekly backup report failed: {e}")

@app.on_event("startup")
async def startup_event():
    """Create database indexes for improved query performance and start schedulers"""
    try:
        # Indexes for jobs collection
        await db.jobs.create_index("license_plate")
        await db.jobs.create_index("status")
        await db.jobs.create_index("created_at")
        await db.jobs.create_index("authority_id")
        await db.jobs.create_index("assigned_service_id")
        await db.jobs.create_index("job_number")
        await db.jobs.create_index([("license_plate", 1), ("status", 1)])  # Compound index
        
        # Polling performance indices
        await db.jobs.create_index([("authority_id", 1), ("updated_at", 1)])
        await db.jobs.create_index([("created_by_id", 1), ("updated_at", 1)])
        await db.jobs.create_index([("assigned_service_id", 1), ("updated_at", 1)])
        
        # DSGVO cleanup index
        await db.jobs.create_index([("status", 1), ("released_at", 1), ("anonymized", 1)])
        
        # Indexes for users collection
        await db.users.create_index("email", unique=True)
        await db.users.create_index("role")
        await db.users.create_index("linked_authorities")
        
        # Indexes for audit_logs collection
        await db.audit_logs.create_index("timestamp")
        await db.audit_logs.create_index("action")
        await db.audit_logs.create_index("user_id")
        
        # Indexes for backup_jobs collection
        await db.backup_jobs.create_index("backup_type")
        await db.backup_jobs.create_index("status")
        await db.backup_jobs.create_index("created_at")
        await db.backup_jobs.create_index("retention_class")
        
        logger.info("Database indexes created successfully")
        
        # Start DSGVO scheduler - runs every day at 03:00 AM
        scheduler.add_job(
            dsgvo_data_cleanup,
            CronTrigger(hour=3, minute=0),
            id="dsgvo_cleanup",
            name="DSGVO Daten-Anonymisierung",
            replace_existing=True
        )
        
        # Start Backup schedulers
        # Täglich um 02:00 Uhr - Datenbank-Backup
        scheduler.add_job(
            scheduled_database_backup,
            CronTrigger(hour=2, minute=0),
            id="daily_db_backup",
            name="Tägliches Datenbank-Backup",
            replace_existing=True
        )
        
        # Täglich um 02:30 Uhr - Storage-Backup
        scheduler.add_job(
            scheduled_storage_backup,
            CronTrigger(hour=2, minute=30),
            id="daily_storage_backup",
            name="Tägliches Storage-Backup",
            replace_existing=True
        )
        
        # Am 1. jedes Monats um 01:00 Uhr - Monatliches Komplett-Backup
        scheduler.add_job(
            scheduled_monthly_backup,
            CronTrigger(day=1, hour=1, minute=0),
            id="monthly_backup",
            name="Monatliches Komplett-Backup",
            replace_existing=True
        )
        
        # Täglich um 04:00 Uhr - Aufbewahrungsregeln anwenden
        scheduler.add_job(
            scheduled_retention_cleanup,
            CronTrigger(hour=4, minute=0),
            id="retention_cleanup",
            name="Backup-Aufbewahrung Aufräumen",
            replace_existing=True
        )
        
        # Jeden Montag um 08:00 Uhr - Wöchentlicher Backup-Report
        scheduler.add_job(
            scheduled_weekly_backup_report,
            CronTrigger(day_of_week='mon', hour=8, minute=0),
            id="weekly_backup_report",
            name="Wöchentlicher Backup-Report",
            replace_existing=True
        )
        
        scheduler.start()
        logger.info(f"DSGVO Scheduler gestartet - Retention: {DSGVO_RETENTION_DAYS} Tage")
        logger.info("Backup Scheduler gestartet - Tägliche Backups um 02:00/02:30 Uhr, Monatliche am 1.")
        logger.info("Wöchentlicher Backup-Report: Jeden Montag um 08:00 Uhr")
        
    except Exception as e:
        logger.error(f"Error during startup: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown(wait=False)
    client.close()
